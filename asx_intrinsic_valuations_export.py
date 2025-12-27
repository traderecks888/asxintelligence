#!/usr/bin/env python3
"""
ASX Intrinsic Valuations + Technicals → Excel/CSV/JSON (+ optional Parquet)
=========================================================================

Outputs
- Full workbook (timestamped) under data/valuations/ASX_Intrinsic_Valuations_YYYYMMDD_HHMMSS.xlsx
- Web-friendly artifacts under --public_dir (default public/data):
    - latest.xlsx   (subset of columns for Cloudflare Pages)
    - latest.csv
    - latest.json
    - latest.parquet (optional if pyarrow/fastparquet installed)
    - manifest.json

Universe refresh
- Uses scrape_asx_universe.py and only refreshes if older than --universe_max_age_days (default 30)

Data sources
- Fundamentals: Yahoo Finance via yfinance (best effort)
- Technicals: 1y daily bars via yfinance history

Run
    python asx_intrinsic_valuations_export.py --resume

Fast test
    python asx_intrinsic_valuations_export.py --limit 50 --sleep 0.15
"""

from __future__ import annotations

import argparse
import re
import json
import math
import os
import subprocess
import zipfile
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Iterable

import numpy as np
import pandas as pd
import yfinance as yf

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------------------
# Global valuation assumptions (aligned to your Core-EnhancedValuationsv5.3.py)
# --------------------------------------------------------------------------------------

ASSUMPTIONS: Dict[str, float] = {
    "risk_free_rate": 0.04,
    "market_risk_premium": 0.06,
    "terminal_growth_rate": 0.02,
    "valuation_period": 5,
    "tax_rate": 0.30,
    "sotp_discount": 0.20,
    "margin_of_safety": 0.25,
    "asset_premium": 0.10,
    "dividend_growth_rate": 0.03,
    "option_years": 5,
    "volatility": 0.30,
}

DEFAULT_GROWTH_FALLBACK = 0.05  # used if revenueGrowth is missing


# --------------------------------------------------------------------------------------
# Excel formatting
# --------------------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def safe_get(d: Dict[str, Any], key: str, default=np.nan):
    try:
        v = d.get(key, default)
        return default if v is None else v
    except Exception:
        return default


def norm_yield(v: Any) -> float:
    """Normalize yield inputs to fraction (0.08 for 8%). If stored as 8.0, convert to 0.08."""
    try:
        x = float(v)
        if np.isnan(x):
            return np.nan
        if abs(x) > 1.5 and abs(x) <= 200:
            return x / 100.0
        return x
    except Exception:
        return np.nan

def epoch_to_date_str(x: Any) -> str:
    """Convert epoch seconds (Yahoo) to YYYY-MM-DD string."""
    try:
        if x is None:
            return ""
        fx = float(x)
        if np.isnan(fx):
            return ""
        return datetime.fromtimestamp(int(fx), tz=ZoneInfo("UTC")).date().isoformat()
    except Exception:
        return ""


def ensure_parent_dir(p: Path) -> None:
    p.parent.mkdir(parents=True, exist_ok=True)


def safe_atomic_write(path: Path, data: str, encoding: str = "utf-8") -> None:
    """Write to a temp file then atomically replace, to avoid partial/corrupt artifacts."""
    ensure_parent_dir(path)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(data, encoding=encoding)
    tmp.replace(path)


def safe_write_json_records(df: pd.DataFrame, path: Path) -> None:
    """Write records JSON and verify it is valid + non-empty before replacing."""
    if df is None or df.shape[0] <= 0:
        raise RuntimeError(f"Refusing to write empty dataset: {path}")

    tmp = path.with_suffix(path.suffix + ".tmp")
    df.to_json(tmp, orient="records")

    # Validate
    try:
        import json as _json
        payload = _json.loads(tmp.read_text(encoding="utf-8"))
        if not isinstance(payload, list) or len(payload) <= 0:
            raise ValueError("records JSON is empty")
    except Exception as e:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        raise RuntimeError(f"Invalid JSON written for {path}: {e}")

    tmp.replace(path)


def parse_iso(dt_str: str) -> Optional[datetime]:
    try:
        return datetime.fromisoformat(dt_str)
    except Exception:
        return None


def is_csv_fresh(csv_path: Path, max_age_days: int) -> bool:
    if not csv_path.exists():
        return False

    # Check last_extracted column if present (mirrors your scraper style)
    try:
        df = pd.read_csv(csv_path)
        if "last_extracted" in df.columns and not df.empty:
            ts = None
            for v in df["last_extracted"].dropna().astype(str).tolist():
                dv = parse_iso(v)
                if dv and (ts is None or dv > ts):
                    ts = dv
            if ts is not None:
                return (datetime.now(ts.tzinfo) if ts.tzinfo else datetime.now()) - ts <= timedelta(days=max_age_days)
    except Exception:
        pass

    try:
        mtime = datetime.fromtimestamp(csv_path.stat().st_mtime)
        return datetime.now() - mtime <= timedelta(days=max_age_days)
    except Exception:
        return False


def run_universe_scraper(scraper_path: Path, output_csv: Path, tickers_txt: Optional[Path], max_age_days: int, source: str, sleep_s: float) -> None:
    cmd = [
        sys.executable,
        str(scraper_path),
        "--output", str(output_csv),
        "--source", source,
        "--sleep", str(sleep_s),
        "--max_age_days", str(max_age_days),
    ]
    if tickers_txt is not None:
        cmd += ["--tickers_txt", str(tickers_txt)]
    print("[info] universe scraper:", " ".join(cmd))
    subprocess.check_call(cmd)


def maybe_refresh_universe(scraper_path: Path, universe_csv: Path, tickers_txt: Optional[Path], max_age_days: int, source: str, sleep_s: float) -> None:
    if is_csv_fresh(universe_csv, max_age_days=max_age_days):
        print(f"[skip] universe CSV looks fresh (<= {max_age_days} days): {universe_csv}")
        return
    print(f"[run] refreshing ASX universe (max_age_days={max_age_days}) -> {universe_csv}")
    run_universe_scraper(scraper_path, universe_csv, tickers_txt, max_age_days=max_age_days, source=source, sleep_s=sleep_s)


def normalize_ticker(code: str) -> str:
    return (code or "").strip().upper().replace(".AX", "")


def yahoo_symbol(code: str) -> str:
    return f"{normalize_ticker(code)}.AX"


# --------------------------------------------------------------------------------------
# GICS Sector normalization (official 11-sector rollup)
# --------------------------------------------------------------------------------------

GICS_11_SECTORS = [
    "Energy",
    "Materials",
    "Industrials",
    "Consumer Discretionary",
    "Consumer Staples",
    "Health Care",
    "Financials",
    "Information Technology",
    "Communication Services",
    "Utilities",
    "Real Estate",
]

_GICS_GARBAGE = {
    "", "na", "n/a", "none", "null", "unknown",
    "not applicable", "not applic", "class pend", "class pending",
}

_GICS_ALIASES = [
    ("energy", "Energy"),
    ("materials", "Materials"),
    ("basic materials", "Materials"),
    ("metals", "Materials"),
    ("mining", "Materials"),

    ("industrials", "Industrials"),
    ("industrial", "Industrials"),
    ("capital goods", "Industrials"),
    ("transportation", "Industrials"),

    ("consumer discretionary", "Consumer Discretionary"),
    ("consumer cyclical", "Consumer Discretionary"),
    ("cyclical", "Consumer Discretionary"),

    ("consumer staples", "Consumer Staples"),
    ("consumer defensive", "Consumer Staples"),
    ("defensive", "Consumer Staples"),

    ("health care", "Health Care"),
    ("healthcare", "Health Care"),
    ("biotechnology", "Health Care"),
    ("biotech", "Health Care"),
    ("pharmaceutical", "Health Care"),

    ("financials", "Financials"),
    ("financial", "Financials"),
    ("financial services", "Financials"),
    ("banks", "Financials"),
    ("insurance", "Financials"),

    ("information technology", "Information Technology"),
    ("technology", "Information Technology"),
    ("tech", "Information Technology"),
    ("software", "Information Technology"),

    ("communication services", "Communication Services"),
    ("communication", "Communication Services"),
    ("telecom", "Communication Services"),
    ("telecommunications", "Communication Services"),
    ("media", "Communication Services"),

    ("utilities", "Utilities"),
    ("utility", "Utilities"),

    ("real estate", "Real Estate"),
    ("reit", "Real Estate"),
]


def _clean_classification(x: str) -> str:
    s = str(x or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^a-z0-9\s&/-]", "", s)
    return s.strip()


def normalize_gics_sector(sector: str, industry: str = "") -> str:
    """Deterministic roll-up to the official 11 GICS sectors.

    Inputs are the existing classification strings (often Yahoo/yfinance-like). We never "invent"
    new categories—only map to the 11 sectors or return "Unknown".
    """
    a = _clean_classification(sector)
    b = _clean_classification(industry)
    if a in _GICS_GARBAGE and b in _GICS_GARBAGE:
        return "Unknown"

    # Exact match / contains mapping (sector first, then industry)
    for needle, out in _GICS_ALIASES:
        if needle in a or needle in b:
            return out

    # A few high-signal fallbacks (industry-only)
    if any(k in b for k in ["bank", "insurance", "asset management", "capital markets"]):
        return "Financials"
    if any(k in b for k in ["oil", "gas", "coal", "uranium", "energy"]):
        return "Energy"
    if any(k in b for k in ["reit", "property", "real estate"]):
        return "Real Estate"

    return "Unknown"


# Normal CDF without SciPy: N(x) = 0.5*(1 + erf(x/sqrt(2)))
def norm_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def black_scholes_call(S: float, K: float, T: float, r: float, sigma: float) -> float:
    if S <= 0 or K <= 0 or T <= 0 or sigma <= 0:
        return float("nan")
    d1 = (math.log(S / K) + (r + (sigma ** 2) / 2.0) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)
    return S * norm_cdf(d1) - K * math.exp(-r * T) * norm_cdf(d2)


def try_get_info(t: yf.Ticker) -> Dict[str, Any]:
    try:
        return t.get_info()
    except Exception:
        try:
            return t.info
        except Exception:
            return {}


def get_balance_sheet_items(t: yf.Ticker) -> Dict[str, float]:
    out = {
        "cash": np.nan,
        "total_assets": np.nan,
        "total_liabilities": np.nan,
        "net_tangible_assets": np.nan,
    }

    def _probe(sheet) -> bool:
        if sheet is None:
            return False
        try:
            idx = getattr(sheet, "index", [])
            if idx is None:
                return False

            for item in ["Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments"]:
                if item in idx:
                    out["cash"] = float(sheet.loc[item].iloc[0])
                    break

            if "Total Assets" in idx:
                out["total_assets"] = float(sheet.loc["Total Assets"].iloc[0])

            for item in ["Total Liabilities Net Minority Interest", "Total Liabilities"]:
                if item in idx:
                    out["total_liabilities"] = float(sheet.loc[item].iloc[0])
                    break

            if "Net Tangible Assets" in idx:
                out["net_tangible_assets"] = float(sheet.loc["Net Tangible Assets"].iloc[0])
            elif not np.isnan(out["total_assets"]) and not np.isnan(out["total_liabilities"]):
                out["net_tangible_assets"] = out["total_assets"] - out["total_liabilities"]

            return not all(np.isnan(v) for v in out.values())
        except Exception:
            return False

    if _probe(getattr(t, "quarterly_balance_sheet", None)):
        return out
    _probe(getattr(t, "balance_sheet", None))
    return out


def clip_growth(g: float) -> float:
    if np.isnan(g):
        return DEFAULT_GROWTH_FALLBACK
    return float(max(-0.10, min(0.25, g)))


def calc_valuations(row: Dict[str, Any]) -> Dict[str, Any]:
    res: Dict[str, Any] = {}

    fcf = float(safe_get(row, "freeCashflow", 0.0) or 0.0)
    growth_raw = safe_get(row, "revenueGrowth", np.nan)
    growth = clip_growth(float(growth_raw) if not np.isnan(growth_raw) else np.nan)
    beta = float(safe_get(row, "beta", 1.0) or 1.0)

    total_debt = safe_get(row, "totalDebt", 0.0)
    total_debt = float(total_debt) if not np.isnan(total_debt) else 0.0

    cash = safe_get(row, "cash", np.nan)
    cash = float(cash) if not np.isnan(cash) else 0.0

    shares_out = safe_get(row, "sharesOutstanding", np.nan)
    shares_out = float(shares_out) if not np.isnan(shares_out) else np.nan

    current_price = safe_get(row, "currentPrice", np.nan)
    current_price = float(current_price) if not np.isnan(current_price) else np.nan

    book_value_ps_yahoo = safe_get(row, "bookValue", np.nan)  # per share
    book_value_ps_yahoo = float(book_value_ps_yahoo) if not np.isnan(book_value_ps_yahoo) else np.nan

    roe = safe_get(row, "returnOnEquity", np.nan)
    roe = float(roe) if not np.isnan(roe) else np.nan

    total_revenue = safe_get(row, "totalRevenue", np.nan)
    total_revenue = float(total_revenue) if not np.isnan(total_revenue) else np.nan

    operating_margin = safe_get(row, "operatingMargins", np.nan)
    operating_margin = float(operating_margin) if not np.isnan(operating_margin) else np.nan

    ebitda = safe_get(row, "ebitda", np.nan)
    ebitda = float(ebitda) if not np.isnan(ebitda) else np.nan

    dividend_rate = safe_get(row, "dividendRate", np.nan)
    dividend_rate = float(dividend_rate) if not np.isnan(dividend_rate) else np.nan

    dividend_yield = safe_get(row, "dividendYield", np.nan)
    dividend_yield = float(dividend_yield) if not np.isnan(dividend_yield) else np.nan

    trailing_eps = safe_get(row, "trailingEps", np.nan)
    trailing_eps = float(trailing_eps) if not np.isnan(trailing_eps) else np.nan

    earnings_growth = safe_get(row, "earningsGrowth", np.nan)
    earnings_growth = float(earnings_growth) if not np.isnan(earnings_growth) else np.nan

    rf = float(ASSUMPTIONS["risk_free_rate"])
    mrp = float(ASSUMPTIONS["market_risk_premium"])
    tgr = float(ASSUMPTIONS["terminal_growth_rate"])
    years = int(ASSUMPTIONS["valuation_period"])
    tax = float(ASSUMPTIONS["tax_rate"])

    cost_of_equity = rf + beta * mrp
    wacc = cost_of_equity

    res["growth_input_used"] = growth
    res["cost_of_equity"] = cost_of_equity
    res["wacc"] = wacc

    dcf_price = np.nan
    enterprise_value = np.nan
    equity_value = np.nan
    net_debt = np.nan
    dcf_estimate = ""

    if not np.isnan(shares_out) and shares_out > 0:
        try:
            projected_fcf = [fcf * ((1.0 + growth) ** y) for y in range(1, years + 1)]
            denom = max(0.01, (wacc - tgr))
            terminal_value = (projected_fcf[-1] * (1.0 + tgr)) / denom

            discounted_cf = [cf / ((1.0 + wacc) ** y) for y, cf in enumerate(projected_fcf, 1)]
            discounted_tv = terminal_value / ((1.0 + wacc) ** years)

            enterprise_value = float(sum(discounted_cf) + discounted_tv)
            net_debt = float(total_debt - cash)
            equity_value = float(enterprise_value - net_debt)
            dcf_price = float(equity_value / shares_out)

            if not np.isnan(current_price) and not np.isnan(dcf_price):
                dcf_estimate = "Undervalued" if dcf_price > current_price else "Overvalued"
        except Exception:
            pass

    res.update({
        "enterprise_value_dcf": enterprise_value,
        "equity_value_dcf": equity_value,
        "net_debt": net_debt,
        "dcf_price": dcf_price,
        "dcf_estimate": dcf_estimate,
    })

    ddm_price = np.nan
    payout_ratio = np.nan
    try:
        g_div = float(ASSUMPTIONS["dividend_growth_rate"])
        if not np.isnan(dividend_rate) and dividend_rate > 0 and (cost_of_equity - g_div) > 0:
            ddm_price = float(dividend_rate / (cost_of_equity - g_div))
        if not np.isnan(dividend_rate) and not np.isnan(trailing_eps) and trailing_eps > 0:
            payout_ratio = float(dividend_rate / trailing_eps)
    except Exception:
        pass

    res.update({
        "ddm_price": ddm_price,
        "dividend_rate": dividend_rate,
        "dividend_yield": dividend_yield,
        "payout_ratio": payout_ratio,
        "last_dividend_value": safe_get(row, "lastDividendValue", np.nan),
        "last_dividend_date": safe_get(row, "lastDividendDate", np.nan),
    })

    epv_price = np.nan
    normalized_earnings = np.nan
    epv_ebit_multiple = np.nan
    try:
        if not np.isnan(total_revenue) and not np.isnan(operating_margin):
            normalized_earnings = float(total_revenue * operating_margin * (1.0 - tax))
            epv = float(normalized_earnings / cost_of_equity) if cost_of_equity > 0 else np.nan
            if not np.isnan(shares_out) and shares_out > 0:
                epv_price = float((epv + cash - total_debt) / shares_out)
            if not np.isnan(ebitda) and ebitda > 0 and not np.isnan(epv):
                epv_ebit_multiple = float(epv / ebitda)
    except Exception:
        pass

    res.update({"epv_price": epv_price, "normalized_earnings": normalized_earnings, "epv_ebit_multiple": epv_ebit_multiple})

    ri_price = np.nan
    try:
        if not np.isnan(book_value_ps_yahoo) and not np.isnan(roe):
            residual_income = book_value_ps_yahoo * (roe - cost_of_equity)
            ri_price = float(book_value_ps_yahoo + (residual_income / (1.0 + cost_of_equity)))
    except Exception:
        pass
    res["residual_income_price"] = ri_price

    asset_price = np.nan
    try:
        nta = safe_get(row, "net_tangible_assets", np.nan)
        if not np.isnan(nta) and not np.isnan(shares_out) and shares_out > 0:
            asset_price = float((float(nta) * (1.0 + float(ASSUMPTIONS["asset_premium"]))) / shares_out)
    except Exception:
        pass
    res["asset_based_price"] = asset_price

    sotp_price = np.nan
    try:
        if not np.isnan(dcf_price):
            sotp_price = float(dcf_price * (1.0 - float(ASSUMPTIONS["sotp_discount"])))
    except Exception:
        pass
    res["sotp_price"] = sotp_price

    implied_g = np.nan
    try:
        if not np.isnan(current_price) and current_price > 0 and not np.isnan(shares_out) and shares_out > 0 and fcf > 0:
            target_equity = current_price * shares_out
            target_ev = target_equity + total_debt - cash
            low, high = -0.10, 0.20
            tolerance = 0.001

            for _ in range(100):
                mid = (low + high) / 2.0
                projected_fcf = [fcf * ((1.0 + mid) ** y) for y in range(1, years + 1)]
                denom = max(0.01, (wacc - tgr))
                terminal_value = (projected_fcf[-1] * (1.0 + tgr)) / denom
                calc_ev = sum([cf / ((1.0 + wacc) ** y) for y, cf in enumerate(projected_fcf, 1)]) + (terminal_value / ((1.0 + wacc) ** years))

                if calc_ev > target_ev:
                    high = mid
                else:
                    low = mid

                if abs(calc_ev - target_ev) < tolerance:
                    break

            implied_g = float(mid)
    except Exception:
        pass
    res["reverse_dcf_implied_growth"] = implied_g

    opt_value = np.nan
    strike = np.nan
    try:
        if not np.isnan(current_price) and current_price > 0:
            strike = float(current_price)
            opt_value = float(black_scholes_call(
                S=float(current_price),
                K=strike,
                T=float(ASSUMPTIONS["option_years"]),
                r=float(ASSUMPTIONS["risk_free_rate"]),
                sigma=float(ASSUMPTIONS["volatility"]),
            ))
    except Exception:
        pass
    res.update({"option_value": opt_value, "option_strike": strike, "option_volatility": float(ASSUMPTIONS["volatility"])})

    peg = np.nan
    try:
        trailing_pe = safe_get(row, "trailingPE", np.nan)
        if not np.isnan(trailing_pe) and not np.isnan(earnings_growth) and earnings_growth != 0:
            peg = float(trailing_pe / (earnings_growth * 100.0))
    except Exception:
        pass
    res["peg_ratio"] = peg

    def _prem(v: float) -> float:
        if np.isnan(current_price) or current_price <= 0 or np.isnan(v):
            return np.nan
        return float((v - current_price) / current_price)

    res.update({
        "dcf_prem": _prem(dcf_price),
        "ri_prem": _prem(ri_price),
        "asset_prem": _prem(asset_price),
        "sotp_prem": _prem(sotp_price),
        "ddm_prem": _prem(ddm_price),
        "epv_prem": _prem(epv_price),
        "opt_prem": _prem(opt_value),
        "margin_of_safety": float(ASSUMPTIONS["margin_of_safety"]),
        "mos_buy_price": float(dcf_price * (1.0 - float(ASSUMPTIONS["margin_of_safety"]))) if not np.isnan(dcf_price) else np.nan,
    })

    prem_cols = ["dcf_prem", "ri_prem", "asset_prem", "sotp_prem", "ddm_prem"]
    res["undervalued_methods_count"] = int(sum(1 for c in prem_cols if not np.isnan(res.get(c, np.nan)) and float(res[c]) > 0))

    return res


def compute_extra_metrics(info: Dict[str, Any], bs: Dict[str, float], shares_out: float) -> Dict[str, Any]:
    out: Dict[str, Any] = {}

    market_cap = safe_get(info, "marketCap", np.nan)
    fcf = safe_get(info, "freeCashflow", np.nan)
    out["fcf_yield"] = (float(fcf) / float(market_cap)) if (not np.isnan(market_cap) and market_cap and not np.isnan(fcf)) else np.nan

    book_ps = safe_get(info, "bookValue", np.nan)  # Yahoo "bookValue" is per share
    out["book_value_per_share_yahoo"] = float(book_ps) if not np.isnan(book_ps) else np.nan
    out["book_value_total_yahoo"] = (float(book_ps) * float(shares_out)) if (not np.isnan(book_ps) and not np.isnan(shares_out) and shares_out > 0) else np.nan

    equity_bs = np.nan
    if not np.isnan(bs.get("total_assets", np.nan)) and not np.isnan(bs.get("total_liabilities", np.nan)):
        equity_bs = float(bs["total_assets"]) - float(bs["total_liabilities"])
    out["book_value_total_bs_equity"] = equity_bs
    out["book_value_per_share_bs_equity"] = (float(equity_bs) / float(shares_out)) if (not np.isnan(equity_bs) and not np.isnan(shares_out) and shares_out > 0) else np.nan

    nta = bs.get("net_tangible_assets", np.nan)
    out["nta_total"] = float(nta) if not np.isnan(nta) else np.nan
    out["nta_per_share"] = (float(nta) / float(shares_out)) if (not np.isnan(nta) and not np.isnan(shares_out) and shares_out > 0) else np.nan

    out["debt_to_equity"] = safe_get(info, "debtToEquity", np.nan)
    out["current_ratio"] = safe_get(info, "currentRatio", np.nan)
    out["quick_ratio"] = safe_get(info, "quickRatio", np.nan)

    out["gross_margin"] = safe_get(info, "grossMargins", np.nan)
    out["operating_margin"] = safe_get(info, "operatingMargins", np.nan)
    out["profit_margin"] = safe_get(info, "profitMargins", np.nan)

    out["roa"] = safe_get(info, "returnOnAssets", np.nan)
    out["roe"] = safe_get(info, "returnOnEquity", np.nan)

    out["trailing_pe"] = safe_get(info, "trailingPE", np.nan)
    out["forward_pe"] = safe_get(info, "forwardPE", np.nan)
    out["price_to_book"] = safe_get(info, "priceToBook", np.nan)
    out["ev_to_ebitda"] = safe_get(info, "enterpriseToEbitda", np.nan)

    total_debt = safe_get(info, "totalDebt", np.nan)
    cash = bs.get("cash", np.nan)
    ebitda = safe_get(info, "ebitda", np.nan)
    out["net_debt_calc"] = (float(total_debt) - float(cash)) if (not np.isnan(total_debt) and not np.isnan(cash)) else np.nan
    out["net_debt_to_ebitda"] = (float(out["net_debt_calc"]) / float(ebitda)) if (not np.isnan(out["net_debt_calc"]) and not np.isnan(ebitda) and ebitda) else np.nan

    out["held_pct_insiders"] = safe_get(info, "heldPercentInsiders", np.nan)
    out["held_pct_institutions"] = safe_get(info, "heldPercentInstitutions", np.nan)
    out["short_pct_float"] = safe_get(info, "shortPercentOfFloat", np.nan)

    enterprise_value = safe_get(info, "enterpriseValue", np.nan)
    if np.isnan(enterprise_value) and not np.isnan(market_cap):
        out["enterprise_value_yahoo_or_calc"] = float(market_cap) + (float(total_debt) if not np.isnan(total_debt) else 0.0) - (float(cash) if not np.isnan(cash) else 0.0)
    else:
        out["enterprise_value_yahoo_or_calc"] = enterprise_value

    return out


def _rsi(close: pd.Series, period: int = 14) -> float:
    if close is None or len(close) < period + 2:
        return float("nan")
    delta = close.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)

    # Wilder's smoothing via EMA with alpha=1/period
    avg_gain = gain.ewm(alpha=1/period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1/period, adjust=False).mean()

    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return float(rsi.iloc[-1])


def _atr(df: pd.DataFrame, period: int = 14) -> float:
    if df is None or df.empty or len(df) < period + 2:
        return float("nan")
    high = df["High"].astype(float)
    low = df["Low"].astype(float)
    close = df["Close"].astype(float)

    prev_close = close.shift(1)
    tr = pd.concat([
        (high - low),
        (high - prev_close).abs(),
        (low - prev_close).abs()
    ], axis=1).max(axis=1)

    atr = tr.ewm(alpha=1/period, adjust=False).mean()
    return float(atr.iloc[-1])




# --------------------------------------------------------------------------------------
# Extra technical helpers (trend + SR levels)
# --------------------------------------------------------------------------------------

def _ema(series: pd.Series, span: int) -> pd.Series:
    return series.ewm(span=span, adjust=False).mean()


def _macd(close: pd.Series, fast: int = 12, slow: int = 26, signal: int = 9) -> Tuple[float, float, float]:
    if close is None or close.dropna().empty:
        return (np.nan, np.nan, np.nan)
    ema_fast = _ema(close, fast)
    ema_slow = _ema(close, slow)
    macd = ema_fast - ema_slow
    sig = _ema(macd, signal)
    hist = macd - sig
    return (float(macd.iloc[-1]), float(sig.iloc[-1]), float(hist.iloc[-1]))


def _adx(df: pd.DataFrame, n: int = 14) -> float:
    """Average Directional Index (ADX). Higher ~ stronger trend (regardless of direction)."""
    if df is None or df.empty:
        return np.nan
    for c in ("High","Low","Close"):
        if c not in df.columns:
            return np.nan
    high = df["High"].astype(float)
    low = df["Low"].astype(float)
    close = df["Close"].astype(float)

    up_move = high.diff()
    down_move = -low.diff()

    plus_dm = np.where((up_move > down_move) & (up_move > 0), up_move, 0.0)
    minus_dm = np.where((down_move > up_move) & (down_move > 0), down_move, 0.0)

    tr = pd.concat([
        (high - low),
        (high - close.shift(1)).abs(),
        (low - close.shift(1)).abs()
    ], axis=1).max(axis=1)

    if tr.dropna().empty:
        return np.nan

    atr = tr.ewm(alpha=1/n, adjust=False).mean()
    plus_di = 100 * pd.Series(plus_dm, index=df.index).ewm(alpha=1/n, adjust=False).mean() / atr
    minus_di = 100 * pd.Series(minus_dm, index=df.index).ewm(alpha=1/n, adjust=False).mean() / atr

    dx = (100 * (plus_di - minus_di).abs() / (plus_di + minus_di)).replace([np.inf, -np.inf], np.nan)
    adx = dx.ewm(alpha=1/n, adjust=False).mean()
    return float(adx.iloc[-1]) if not adx.dropna().empty else np.nan


def _stoch(df: pd.DataFrame, k_period: int = 14, d_period: int = 3) -> Tuple[float, float]:
    if df is None or df.empty:
        return (np.nan, np.nan)
    for c in ("High","Low","Close"):
        if c not in df.columns:
            return (np.nan, np.nan)
    high = df["High"].astype(float)
    low = df["Low"].astype(float)
    close = df["Close"].astype(float)

    lo = low.rolling(k_period).min()
    hi = high.rolling(k_period).max()
    denom = (hi - lo).replace(0, np.nan)
    k = 100 * (close - lo) / denom
    d = k.rolling(d_period).mean()
    return (float(k.iloc[-1]) if not k.dropna().empty else np.nan,
            float(d.iloc[-1]) if not d.dropna().empty else np.nan)


def _bollinger(close: pd.Series, n: int = 20, k: float = 2.0) -> Tuple[float, float]:
    """Returns (%B, bandwidth)."""
    if close is None or close.dropna().empty:
        return (np.nan, np.nan)
    mid = close.rolling(n).mean()
    sd = close.rolling(n).std(ddof=0)
    upper = mid + k * sd
    lower = mid - k * sd
    width = (upper - lower) / mid
    pb = (close - lower) / (upper - lower)
    pb = pb.replace([np.inf, -np.inf], np.nan)
    width = width.replace([np.inf, -np.inf], np.nan)
    return (float(pb.iloc[-1]) if not pb.dropna().empty else np.nan,
            float(width.iloc[-1]) if not width.dropna().empty else np.nan)


def _pivot_levels(df: pd.DataFrame, win: int = 3) -> List[float]:
    """Simple pivot-based levels from local highs/lows."""
    if df is None or df.empty:
        return []
    if "High" not in df.columns or "Low" not in df.columns:
        return []
    high = df["High"].astype(float)
    low = df["Low"].astype(float)
    w = 2 * win + 1
    ph = high[(high == high.rolling(w, center=True).max())]
    pl = low[(low == low.rolling(w, center=True).min())]
    levels = pd.concat([ph, pl]).dropna().astype(float).tolist()
    return levels


def _cluster_levels(levels: List[float], merge_pct: float = 0.005) -> List[float]:
    """Merge nearby levels (within merge_pct) to reduce duplicates."""
    lv = [float(x) for x in levels if x is not None and np.isfinite(x)]
    if not lv:
        return []
    lv.sort()
    clusters: List[List[float]] = [[lv[0]]]
    for x in lv[1:]:
        m = float(sum(clusters[-1]) / len(clusters[-1]))
        if m > 0 and abs(x - m) / m <= merge_pct:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    return [float(sum(c)/len(c)) for c in clusters]


def _nearest_sr(df: pd.DataFrame, px: float, win: int = 3, merge_pct: float = 0.005) -> Dict[str, Any]:
    """Nearest two support + resistance levels, plus % distances and R:R."""
    out: Dict[str, Any] = {
        "S1": np.nan, "S2": np.nan, "R1": np.nan, "R2": np.nan,
        "S1 %": np.nan, "S2 %": np.nan, "R1 %": np.nan, "R2 %": np.nan,
        "R:R": np.nan,
    }
    if not (np.isfinite(px) and px > 0) or df is None or df.empty:
        return out

    levels = _cluster_levels(_pivot_levels(df, win=win), merge_pct=merge_pct)
    if not levels:
        return out

    supports = [x for x in levels if x < px]
    resist = [x for x in levels if x > px]
    supports.sort(reverse=True)
    resist.sort()

    def _pct_support(s): return float((px - s) / px) if np.isfinite(s) and px > 0 else np.nan
    def _pct_res(r): return float((r - px) / px) if np.isfinite(r) and px > 0 else np.nan

    if supports:
        out["S1"] = float(supports[0]); out["S1 %"] = _pct_support(supports[0])
    if len(supports) > 1:
        out["S2"] = float(supports[1]); out["S2 %"] = _pct_support(supports[1])

    if resist:
        out["R1"] = float(resist[0]); out["R1 %"] = _pct_res(resist[0])
    if len(resist) > 1:
        out["R2"] = float(resist[1]); out["R2 %"] = _pct_res(resist[1])

    if np.isfinite(out["R1 %"]) and np.isfinite(out["S1 %"]) and out["S1 %"] > 0:
        out["R:R"] = float(out["R1 %"] / out["S1 %"])
    return out

def compute_technicals(hist: pd.DataFrame, benchmark_rets: Optional[pd.Series] = None, weekly_hist: Optional[pd.DataFrame] = None, monthly_hist: Optional[pd.DataFrame] = None) -> Dict[str, Any]:
    out: Dict[str, Any] = {
        "Return 1m": np.nan,
        "Return 3m": np.nan,
        "Return 6m": np.nan,
        "Return 12m": np.nan,
        "SMA20": np.nan,
        "SMA50": np.nan,
        "SMA200": np.nan,
        "% from SMA20": np.nan,
        "% from SMA50": np.nan,
        "% from SMA200": np.nan,
        "% Dist SMA200D": np.nan,
        "SMA200W": np.nan,
        "% Dist SMA200W": np.nan,
        "MACD (12,26)": np.nan,
        "MACD Signal (9)": np.nan,
        "MACD Hist (12,26,9)": np.nan,
        "ADX14": np.nan,
        "Stoch %K (14)": np.nan,
        "Stoch %D (3)": np.nan,
        "BB %B (20,2)": np.nan,
        "BB Width (20,2)": np.nan,
        # Support/Resistance + R:R
        "Support D1": np.nan, "Support D2": np.nan, "Resistance D1": np.nan, "Resistance D2": np.nan,
        "Support D1 %": np.nan, "Support D2 %": np.nan, "Resistance D1 %": np.nan, "Resistance D2 %": np.nan,
        "R:R (D)": np.nan,
        "Support W1": np.nan, "Support W2": np.nan, "Resistance W1": np.nan, "Resistance W2": np.nan,
        "Support W1 %": np.nan, "Support W2 %": np.nan, "Resistance W1 %": np.nan, "Resistance W2 %": np.nan,
        "R:R (W)": np.nan,
        "Support M1": np.nan, "Support M2": np.nan, "Resistance M1": np.nan, "Resistance M2": np.nan,
        "Support M1 %": np.nan, "Support M2 %": np.nan, "Resistance M1 %": np.nan, "Resistance M2 %": np.nan,
        "R:R (M)": np.nan,

        "52W High": np.nan,
        "52W Low": np.nan,
        "% From 52W High": np.nan,
        "% From 52W Low": np.nan,
        "RSI14": np.nan,
        "ATR (14)": np.nan,
        "ATR% (14)": np.nan,
        "Vol (20d, ann)": np.nan,
        "Vol (60d, ann)": np.nan,
        "Max Drawdown (1y)": np.nan,
        "Avg Vol 20d": np.nan,
        "Avg $Vol 20d": np.nan,
        "Beta vs Benchmark (1y)": np.nan,
    }
    if hist is None or hist.empty or "Close" not in hist.columns:
        return out

    df = hist.dropna(subset=["Close"]).copy()
    close = df["Close"].astype(float)

    # Returns
    def _ret(n: int) -> float:
        if len(close) <= n:
            return float("nan")
        return float(close.iloc[-1] / close.iloc[-(n+1)] - 1.0)

    out["Return 1m"] = _ret(21)
    out["Return 3m"] = _ret(63)
    out["Return 6m"] = _ret(126)
    out["Return 12m"] = _ret(252)

    # Moving averages
    sma20 = close.rolling(20).mean()
    sma50 = close.rolling(50).mean()
    sma200 = close.rolling(200).mean()
    out["SMA20"] = float(sma20.iloc[-1]) if len(sma20.dropna()) else float("nan")
    out["SMA50"] = float(sma50.iloc[-1]) if len(sma50.dropna()) else float("nan")
    out["SMA200"] = float(sma200.iloc[-1]) if len(sma200.dropna()) else float("nan")

    px = float(close.iloc[-1])
    out["% from SMA20"] = float(px / out["SMA20"] - 1.0) if np.isfinite(out["SMA20"]) and out["SMA20"] > 0 else np.nan
    out["% from SMA50"] = float(px / out["SMA50"] - 1.0) if np.isfinite(out["SMA50"]) and out["SMA50"] > 0 else np.nan
    out["% from SMA200"] = float(px / out["SMA200"] - 1.0) if np.isfinite(out["SMA200"]) and out["SMA200"] > 0 else np.nan

    # Alias requested naming (200-day distance)
    out["% Dist SMA200D"] = out["% from SMA200"]

    # 200-week SMA + distance (weekly)
    try:
        if weekly_hist is not None and not weekly_hist.empty and "Close" in weekly_hist.columns:
            wclose = weekly_hist["Close"].astype(float).dropna()
            if len(wclose) >= 200:
                sma200w = float(wclose.rolling(200).mean().iloc[-1])
                out["SMA200W"] = sma200w
                out["% Dist SMA200W"] = float(px / sma200w - 1.0) if np.isfinite(sma200w) and sma200w > 0 else np.nan
    except Exception:
        pass


    # 52-week high/low (use High/Low if available else Close)
    if "High" in df.columns and "Low" in df.columns:
        out["52W High"] = float(df["High"].astype(float).max())
        out["52W Low"] = float(df["Low"].astype(float).min())
    else:
        out["52W High"] = float(close.max())
        out["52W Low"] = float(close.min())

    out["% From 52W High"] = float(px / out["52W High"] - 1.0) if np.isfinite(out["52W High"]) and out["52W High"] > 0 else np.nan
    out["% From 52W Low"] = float(px / out["52W Low"] - 1.0) if np.isfinite(out["52W Low"]) and out["52W Low"] > 0 else np.nan

    # RSI, ATR
    out["RSI14"] = _rsi(close, 14)

    # Extra pro-grade trend/oscillator indicators (lightweight, interpretable)
    try:
        macd, macd_sig, macd_hist = _macd(close, 12, 26, 9)
        out["MACD (12,26)"] = macd
        out["MACD Signal (9)"] = macd_sig
        out["MACD Hist (12,26,9)"] = macd_hist
    except Exception:
        pass

    try:
        out["ADX14"] = _adx(df, 14)
    except Exception:
        pass

    try:
        k, d = _stoch(df, 14, 3)
        out["Stoch %K (14)"] = k
        out["Stoch %D (3)"] = d
    except Exception:
        pass

    try:
        pb, bw = _bollinger(close, 20, 2.0)
        out["BB %B (20,2)"] = pb
        out["BB Width (20,2)"] = bw
    except Exception:
        pass

    atr14 = _atr(df, 14)
    out["ATR (14)"] = atr14
    out["ATR% (14)"] = float(atr14 / px) if np.isfinite(atr14) and px > 0 else np.nan

    # Volatility
    rets = close.pct_change().dropna()
    if len(rets) >= 20:
        out["Vol (20d, ann)"] = float(rets.tail(20).std(ddof=0) * math.sqrt(252))
    if len(rets) >= 60:
        out["Vol (60d, ann)"] = float(rets.tail(60).std(ddof=0) * math.sqrt(252))

    # Max drawdown
    if len(close) >= 2:
        cum = (1 + rets).fillna(0).add(1).cumprod()
        peak = cum.cummax()
        dd = (cum / peak) - 1.0
        out["Max Drawdown (1y)"] = float(dd.min())

    # Liquidity
    if "Volume" in df.columns:
        vol = df["Volume"].astype(float).dropna()
        if len(vol) >= 20:
            out["Avg Vol 20d"] = float(vol.tail(20).mean())
            out["Avg $Vol 20d"] = float(out["Avg Vol 20d"] * close.tail(20).mean())

    # Beta vs benchmark
    if benchmark_rets is not None and len(rets) > 30:
        joined = pd.concat([rets.rename("stock"), benchmark_rets.rename("bench")], axis=1).dropna()
        if len(joined) > 30 and joined["bench"].var(ddof=0) > 0:
            cov = joined.cov(ddof=0).loc["stock", "bench"]
            out["Beta vs Benchmark (1y)"] = float(cov / joined["bench"].var(ddof=0))



    # Support/Resistance + R:R (daily/weekly/monthly)
    try:
        d_df = df.tail(252).copy()
        sr_d = _nearest_sr(d_df, px, win=3, merge_pct=0.006)
        out["Support D1"] = sr_d["S1"]; out["Support D2"] = sr_d["S2"]
        out["Resistance D1"] = sr_d["R1"]; out["Resistance D2"] = sr_d["R2"]
        out["Support D1 %"] = sr_d["S1 %"]; out["Support D2 %"] = sr_d["S2 %"]
        out["Resistance D1 %"] = sr_d["R1 %"]; out["Resistance D2 %"] = sr_d["R2 %"]
        out["R:R (D)"] = sr_d["R:R"]
    except Exception:
        pass

    try:
        if weekly_hist is not None and not weekly_hist.empty:
            w_df = weekly_hist.tail(260).copy()
            sr_w = _nearest_sr(w_df, px, win=2, merge_pct=0.010)
            out["Support W1"] = sr_w["S1"]; out["Support W2"] = sr_w["S2"]
            out["Resistance W1"] = sr_w["R1"]; out["Resistance W2"] = sr_w["R2"]
            out["Support W1 %"] = sr_w["S1 %"]; out["Support W2 %"] = sr_w["S2 %"]
            out["Resistance W1 %"] = sr_w["R1 %"]; out["Resistance W2 %"] = sr_w["R2 %"]
            out["R:R (W)"] = sr_w["R:R"]
    except Exception:
        pass

    try:
        if monthly_hist is not None and not monthly_hist.empty:
            m_df = monthly_hist.tail(180).copy()
            sr_m = _nearest_sr(m_df, px, win=2, merge_pct=0.015)
            out["Support M1"] = sr_m["S1"]; out["Support M2"] = sr_m["S2"]
            out["Resistance M1"] = sr_m["R1"]; out["Resistance M2"] = sr_m["R2"]
            out["Support M1 %"] = sr_m["S1 %"]; out["Support M2 %"] = sr_m["S2 %"]
            out["Resistance M1 %"] = sr_m["R1 %"]; out["Resistance M2 %"] = sr_m["R2 %"]
            out["R:R (M)"] = sr_m["R:R"]
    except Exception:
        pass
    return out


def format_sheet(ws) -> None:
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        width = min(28, max(10, len(str(header)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


@dataclass
class UniverseRow:
    ticker: str
    name: str = ""
    sector: str = ""
    industry: str = ""


def load_universe(universe_csv: Path) -> List[UniverseRow]:
    df = pd.read_csv(universe_csv)
    col_ticker = "ticker" if "ticker" in df.columns else ("Code" if "Code" in df.columns else None)
    if col_ticker is None:
        raise ValueError(f"Universe CSV is missing a ticker column: {universe_csv}")

    def _col(name: str) -> str:
        return name if name in df.columns else ""

    rows: List[UniverseRow] = []
    for _, r in df.iterrows():
        t = normalize_ticker(str(r[col_ticker]))
        if not t:
            continue
        rows.append(UniverseRow(
            ticker=t,
            name=str(r[_col("name")]) if _col("name") else str(r.get("Company name", "")),
            sector=str(r[_col("sector")]) if _col("sector") else str(r.get("GICS industry group", "")),
            industry=str(r[_col("industry")]) if _col("industry") else str(r.get("Industry group", "")),
        ))
    return rows


def maybe_load_materials_energy(roster_csv: Path) -> Optional[pd.DataFrame]:
    if not roster_csv.exists():
        return None
    try:
        df = pd.read_csv(roster_csv)
        if "ticker" not in df.columns:
            return None
        df["ticker"] = df["ticker"].astype(str).map(normalize_ticker)
        keep = [c for c in ["ticker", "products", "stage"] if c in df.columns]
        return df[keep].drop_duplicates("ticker")
    except Exception:
        return None


def fetch_history(symbol: str, period: str, interval: str) -> pd.DataFrame:
    t = yf.Ticker(symbol)
    try:
        h = t.history(period=period, interval=interval, auto_adjust=True)
        return h if isinstance(h, pd.DataFrame) else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def _split_download_frame(df: pd.DataFrame, symbols: List[str]) -> Dict[str, pd.DataFrame]:
    """Split yfinance.download multi-ticker output into per-symbol OHLCV frames.

    yfinance output schema varies by version:
    - MultiIndex columns: (symbol, field)
    - MultiIndex columns: (field, symbol)
    - Single-index columns for a single symbol.
    """
    out: Dict[str, pd.DataFrame] = {}
    if df is None or df.empty:
        return out

    sym_set = set(symbols)

    if isinstance(df.columns, pd.MultiIndex):
        lvl0 = df.columns.get_level_values(0)
        lvl1 = df.columns.get_level_values(1)

        if any(s in sym_set for s in lvl0):
            for s in symbols:
                if s in lvl0:
                    sub = df[s].copy()
                    out[s] = sub
        elif any(s in sym_set for s in lvl1):
            for s in symbols:
                if s in lvl1:
                    sub = df.xs(s, axis=1, level=1, drop_level=True).copy()
                    out[s] = sub
    else:
        # Single-ticker frame
        if len(symbols) == 1:
            out[symbols[0]] = df.copy()

    # Normalize columns
    for s, sub in list(out.items()):
        if sub is None or sub.empty:
            continue
        sub2 = sub.copy()
        if "Adj Close" in sub2.columns and "Close" not in sub2.columns:
            sub2.rename(columns={"Adj Close": "Close"}, inplace=True)
        # Some yfinance versions return lowercase; standardize
        ren = {}
        for c in sub2.columns:
            if isinstance(c, str):
                ren[c] = c[:1].upper() + c[1:]
        if ren:
            sub2.rename(columns=ren, inplace=True)
        out[s] = sub2

    return out


def bulk_fetch_history(symbols: List[str], period: str, interval: str, chunk_size: int = 250) -> Dict[str, pd.DataFrame]:
    """Bulk-download OHLCV for many tickers in chunks.

    This is usually *much* faster than calling Ticker().history per symbol because it reduces
    thousands of HTTP requests to ~N/chunk_size requests.
    """
    # Unique, preserve order
    uniq: List[str] = []
    seen = set()
    for s in symbols:
        s2 = str(s).strip()
        if not s2 or s2 in seen:
            continue
        seen.add(s2)
        uniq.append(s2)

    out: Dict[str, pd.DataFrame] = {}
    if not uniq:
        return out

    chunk_size = int(chunk_size) if chunk_size and int(chunk_size) > 0 else 250

    for i in range(0, len(uniq), chunk_size):
        chunk = uniq[i:i + chunk_size]
        tickers_str = " ".join(chunk)
        try:
            df = yf.download(
                tickers=tickers_str,
                period=period,
                interval=interval,
                group_by="ticker",
                auto_adjust=True,
                threads=True,
                progress=False,
            )
        except Exception as e:
            print(f"[warn] bulk download failed for chunk {i//chunk_size+1}: {e}")
            continue

        out.update(_split_download_frame(df, chunk))

    return out


def data_quality_score(r: Dict[str, Any]) -> float:
    keys = [
        "Price", "Market Cap", "Shares Out", "FCF (Yahoo)",
        "Revenue Growth (Yahoo)", "Beta", "Book Value / Share (Yahoo)",
        "Total Debt", "Cash", "ROE", "DCF Price (5yr)",
        "Vol (20d, ann)", "Avg $Vol 20d",
    ]
    ok = 0
    for k in keys:
        v = r.get(k, np.nan)
        if v is None:
            continue
        if isinstance(v, (int, float)) and np.isnan(v):
            continue
        ok += 1
    return float(ok / len(keys))




def build_out_row(base: Dict[str, Any], tech: Dict[str, Any]) -> Dict[str, Any]:
    """Map internal base fields + computed technicals into the stable, UI-facing row schema."""
    out_row: Dict[str, Any] = {
        "Ticker": base["ticker"],
        "Company": base["company_name"],
        "GICS Sector": base.get("gics_sector", normalize_gics_sector(base.get("sector",""), base.get("industry",""))),
        "Sector": base["sector"],
        "Industry": base["industry"],
        "Currency": base["currency"],
        "As Of": base["asof"],

        "Price": base["currentPrice"],
        "Market Cap": base["marketCap"],
        "Shares Out": base["sharesOutstanding"],

        "Cash": base["cash"],
        "Total Debt": base["totalDebt"],
        "Net Debt (Debt - Cash)": base.get("net_debt", np.nan),
        "Total Assets": base["total_assets"],
        "Total Liabilities": base["total_liabilities"],
        "NTA": base["net_tangible_assets"],
        "NTA / Share": base.get("nta_per_share", np.nan),

        "Book Value (Total, Yahoo)": base.get("book_value_total_yahoo", np.nan),
        "Book Value / Share (Yahoo)": base.get("book_value_per_share_yahoo", np.nan),
        "Book Value (Total, Assets-Liab)": base.get("book_value_total_bs_equity", np.nan),
        "Book Value / Share (Assets-Liab)": base.get("book_value_per_share_bs_equity", np.nan),

        "DCF Price (5yr)": base.get("dcf_price", np.nan),
        "Residual Income Price": base.get("residual_income_price", np.nan),
        "Asset Based Price": base.get("asset_based_price", np.nan),
        "SOTP Price": base.get("sotp_price", np.nan),
        "Dividend Discount Price": base.get("ddm_price", np.nan),
        "Earnings Power Value (EPV) Price": base.get("epv_price", np.nan),
        "Option Pricing Value": base.get("option_value", np.nan),

        "DCF Estimate": base.get("dcf_estimate", ""),

        "DCF Premium/(Discount)": base.get("dcf_prem", np.nan),
        "Residual Income Premium/(Discount)": base.get("ri_prem", np.nan),
        "Asset Based Premium/(Discount)": base.get("asset_prem", np.nan),
        "SOTP Premium/(Discount)": base.get("sotp_prem", np.nan),
        "Dividend Discount Premium/(Discount)": base.get("ddm_prem", np.nan),
        "EPV Premium/(Discount)": base.get("epv_prem", np.nan),
        "Option Pricing Premium/(Discount)": base.get("opt_prem", np.nan),

        "Undervalued Methods Count": base.get("undervalued_methods_count", np.nan),
        "Margin of Safety": base.get("margin_of_safety", np.nan),
        "MOS Buy Price": base.get("mos_buy_price", np.nan),

        "FCF (Yahoo)": base.get("freeCashflow", np.nan),
        "Revenue Growth (Yahoo)": base.get("revenueGrowth", np.nan),
        "Growth Input Used": base.get("growth_input_used", np.nan),
        "Beta": base.get("beta", np.nan),
        "Risk Free Rate": ASSUMPTIONS["risk_free_rate"],
        "Market Risk Premium": ASSUMPTIONS["market_risk_premium"],
        "Cost of Equity": base.get("cost_of_equity", np.nan),
        "WACC": base.get("wacc", np.nan),
        "Terminal Growth Rate": ASSUMPTIONS["terminal_growth_rate"],
        "Valuation Period (yrs)": ASSUMPTIONS["valuation_period"],
        "Tax Rate": ASSUMPTIONS["tax_rate"],

        "FCF Yield": base.get("fcf_yield", np.nan),
        "Trailing PE": base.get("trailing_pe", np.nan),
        "Forward PE": base.get("forward_pe", np.nan),
        "P/B": base.get("price_to_book", np.nan),
        "EV/EBITDA": base.get("ev_to_ebitda", np.nan),
        "Net Debt/EBITDA": base.get("net_debt_to_ebitda", np.nan),

        "ROA": base.get("roa", np.nan),
        "ROE": base.get("roe", np.nan),
        "Gross Margin": base.get("gross_margin", np.nan),
        "Operating Margin": base.get("operating_margin", np.nan),
        "Profit Margin": base.get("profit_margin", np.nan),

        "Debt/Equity": base.get("debt_to_equity", np.nan),
        "Current Ratio": base.get("current_ratio", np.nan),
        "Quick Ratio": base.get("quick_ratio", np.nan),

        "Held % Insiders": base.get("held_pct_insiders", np.nan),
        "Held % Institutions": base.get("held_pct_institutions", np.nan),
        "Short % Float": base.get("short_pct_float", np.nan),
        "Dividend Rate (Yahoo)": base.get("dividend_rate", np.nan),
        "Dividend Yield (Yahoo)": norm_yield(base.get("dividend_yield", np.nan)),

        # Raw Yahoo dividend metadata (as available)
        "Payout Ratio (Yahoo)": base.get("payoutRatio", np.nan),
        "5Y Avg Dividend Yield (Yahoo)": norm_yield(base.get("fiveYearAvgDividendYield", np.nan)),
        "Ex-Dividend Date (Yahoo)": epoch_to_date_str(base.get("exDividendDate", np.nan)),
        "Last Dividend Value (Yahoo)": base.get("lastDividendValue", np.nan),
        "Last Dividend Date (Yahoo)": epoch_to_date_str(base.get("lastDividendDate", np.nan)),

        # Requested calc: trailing annual dividend rate / latest share price
        "Dividend Yield (Latest, Calc)": (
            float(base.get("dividend_rate", np.nan)) / float(base["currentPrice"])
            if (not np.isnan(base.get("dividend_rate", np.nan)) and not np.isnan(base["currentPrice"]) and float(base["currentPrice"]) > 0)
            else np.nan
        ),

        # Delta between Yahoo's dividendYield and the calculated yield above (relative % change)
        "Dividend Yield Δ% (Yahoo→Calc)": (
            ( (float(base.get("dividend_rate", np.nan)) / float(base["currentPrice"])) - float(norm_yield(base.get("dividend_yield", np.nan))) )
            / float(norm_yield(base.get("dividend_yield", np.nan)))
            if (not np.isnan(base.get("dividend_rate", np.nan)) and not np.isnan(base["currentPrice"]) and float(base["currentPrice"]) > 0
                and not np.isnan(base.get("dividend_yield", np.nan)) and float(norm_yield(base.get("dividend_yield", np.nan))) != 0)
            else np.nan
        ),


        "Reverse DCF Implied Growth": base.get("reverse_dcf_implied_growth", np.nan),
        "PEG Ratio": base.get("peg_ratio", np.nan),

        "Enterprise Value (Yahoo/Calc)": base.get("enterprise_value_yahoo_or_calc", np.nan),
    }

    # Add technical columns with the exact names the screener expects
    out_row.update(tech)

    # Add a simple completeness score for quick filtering
    out_row["Data Quality Score"] = data_quality_score(out_row)

    return out_row

def fetch_one(
    ticker_code: str,
    company: UniverseRow,
    benchmark_rets: Optional[pd.Series],
    history_period: str,
    history_interval: str,
    disable_technicals: bool,
    include_long_tf: bool = True,
    hist_override: Optional[pd.DataFrame] = None,
) -> Optional[Dict[str, Any]]:
    sym = yahoo_symbol(ticker_code)
    t = yf.Ticker(sym)
    info = try_get_info(t)
    if not info:
        return None

    # ----------------------------------------------------------------------------------
    # Dividend fallback (Yahoo info fields are often missing for AU tickers)
    # We prefer:
    # 1) info["dividendRate"/"dividendYield"/"lastDividendValue"/"lastDividendDate"] when present
    # 2) yfinance dividends time-series as a fallback
    # ----------------------------------------------------------------------------------
    div_last_value = np.nan
    div_last_date_epoch = np.nan
    div_rate_ttm = np.nan

    try:
        divs = t.dividends  # pandas Series indexed by date
        if isinstance(divs, pd.Series) and not divs.empty:
            div_last_value = float(divs.iloc[-1])
            try:
                last_dt = pd.to_datetime(divs.index[-1]).to_pydatetime()
                # store epoch seconds (UTC) to align with existing handling
                div_last_date_epoch = float(int(last_dt.replace(tzinfo=ZoneInfo("UTC")).timestamp()))
            except Exception:
                div_last_date_epoch = np.nan

            try:
                cutoff = pd.Timestamp.utcnow().tz_localize("UTC") - pd.Timedelta(days=365)
                div_rate_ttm = float(divs[divs.index >= cutoff].sum())
            except Exception:
                div_rate_ttm = np.nan
    except Exception:
        pass

    bs = get_balance_sheet_items(t)

    current_price = safe_get(info, "currentPrice", np.nan)
    market_cap = safe_get(info, "marketCap", np.nan)

    shares_out = safe_get(info, "sharesOutstanding", np.nan)
    if (np.isnan(shares_out) or shares_out <= 0) and not np.isnan(market_cap) and not np.isnan(current_price) and current_price:
        shares_out = float(market_cap) / float(current_price)

    base: Dict[str, Any] = {
        "ticker": ticker_code,
        "company_name": company.name or safe_get(info, "shortName", ""),
        "sector": company.sector,
        "industry": company.industry,
        "currency": safe_get(info, "currency", "AUD"),
        "asof": datetime.now().strftime("%Y-%m-%d"),
        "currentPrice": current_price,
        "marketCap": market_cap,
        # Dividend fields (prefer Yahoo info; fallback to dividends series where missing)
        "dividendRate": safe_get(info, "dividendRate", np.nan),
        "payoutRatio": safe_get(info, "payoutRatio", np.nan),
        "exDividendDate": safe_get(info, "exDividendDate", np.nan),
        "fiveYearAvgDividendYield": safe_get(info, "fiveYearAvgDividendYield", np.nan),
        "dividendYield": safe_get(info, "dividendYield", np.nan),
        "lastDividendValue": safe_get(info, "lastDividendValue", np.nan),
        "lastDividendDate": safe_get(info, "lastDividendDate", np.nan),
        "sharesOutstanding": shares_out,
        "beta": safe_get(info, "beta", np.nan),
        "freeCashflow": safe_get(info, "freeCashflow", np.nan),
        "revenueGrowth": safe_get(info, "revenueGrowth", np.nan),
        "earningsGrowth": safe_get(info, "earningsGrowth", np.nan),
        "totalDebt": safe_get(info, "totalDebt", np.nan),
        "bookValue": safe_get(info, "bookValue", np.nan),  # per share
        "returnOnEquity": safe_get(info, "returnOnEquity", np.nan),
        "trailingPE": safe_get(info, "trailingPE", np.nan),
        "enterpriseToEbitda": safe_get(info, "enterpriseToEbitda", np.nan),
        "priceToBook": safe_get(info, "priceToBook", np.nan),
        "dividendYield": safe_get(info, "dividendYield", np.nan),
        "dividendRate": safe_get(info, "dividendRate", np.nan),
        "trailingEps": safe_get(info, "trailingEps", np.nan),
        "totalRevenue": safe_get(info, "totalRevenue", np.nan),
        "ebitda": safe_get(info, "ebitda", np.nan),
        "operatingMargins": safe_get(info, "operatingMargins", np.nan),
        "cash": bs["cash"],
        "total_assets": bs["total_assets"],
        "total_liabilities": bs["total_liabilities"],
        "net_tangible_assets": bs["net_tangible_assets"],
    }

    # Deterministic official GICS roll-up (for consistent filtering / RRG match)
    base["gics_sector"] = normalize_gics_sector(base.get("sector", ""), base.get("industry", ""))


    # Fill missing dividend fields from fallback series
    try:
        if np.isnan(safe_get(base, "lastDividendValue", np.nan)) and not np.isnan(div_last_value):
            base["lastDividendValue"] = div_last_value
        if np.isnan(safe_get(base, "lastDividendDate", np.nan)) and not np.isnan(div_last_date_epoch):
            base["lastDividendDate"] = div_last_date_epoch
        if np.isnan(safe_get(base, "dividendRate", np.nan)) and not np.isnan(div_rate_ttm):
            base["dividendRate"] = div_rate_ttm
        # dividendYield can be derived if missing
        if np.isnan(safe_get(base, "dividendYield", np.nan)) and not np.isnan(safe_get(base, "dividendRate", np.nan)) and not np.isnan(current_price) and current_price:
            base["dividendYield"] = float(base["dividendRate"]) / float(current_price)
    except Exception:
        pass

    base.update(calc_valuations(base))
    extras = compute_extra_metrics(info, bs, float(base["sharesOutstanding"]) if not np.isnan(base["sharesOutstanding"]) else np.nan)
    base.update(extras)

    # Technicals (daily + weekly/monthly for longer-horizon signals)
    tech: Dict[str, Any] = {}
    if not disable_technicals:
        hist = hist_override if (isinstance(hist_override, pd.DataFrame) and not hist_override.empty) else fetch_history(sym, history_period, history_interval)
        hist_w = None
        hist_m = None
        if include_long_tf:
            # Weekly/monthly are small row-count and unlock 200-week SMA + multi-timeframe SR
            hist_w = fetch_history(sym, "5y", "1wk")
            hist_m = fetch_history(sym, "10y", "1mo")
        tech = compute_technicals(hist, benchmark_rets=benchmark_rets, weekly_hist=hist_w, monthly_hist=hist_m)
    else:
        tech = compute_technicals(pd.DataFrame(), benchmark_rets=None, weekly_hist=None, monthly_hist=None)


    # ----------------------------------------------------------------------------------
    # Dividend & ownership extras (for screening / income overlays)
    # ----------------------------------------------------------------------------------
    last_div = base.get("last_dividend_value", np.nan)
    try:
        last_div = float(last_div) if not np.isnan(last_div) else np.nan
    except Exception:
        last_div = np.nan

    last_div_date = base.get("last_dividend_date", np.nan)

    # Output row (stable user-facing columns)
    out_row: Dict[str, Any] = {
        "Ticker": base["ticker"],
        "Company": base["company_name"],
        "GICS Sector": base.get("gics_sector", normalize_gics_sector(base.get("sector",""), base.get("industry",""))),
        "Sector": base["sector"],
        "Industry": base["industry"],
        "Currency": base["currency"],
        "As Of": base["asof"],

        "Price": base["currentPrice"],
        "Market Cap": base["marketCap"],
        "Shares Out": base["sharesOutstanding"],

        "Cash": base["cash"],
        "Total Debt": base["totalDebt"],
        "Net Debt (Debt - Cash)": base.get("net_debt", np.nan),
        "Total Assets": base["total_assets"],
        "Total Liabilities": base["total_liabilities"],
        "NTA": base["net_tangible_assets"],
        "NTA / Share": base.get("nta_per_share", np.nan),

        "Book Value (Total, Yahoo)": base.get("book_value_total_yahoo", np.nan),
        "Book Value / Share (Yahoo)": base.get("book_value_per_share_yahoo", np.nan),
        "Book Value (Total, Assets-Liab)": base.get("book_value_total_bs_equity", np.nan),
        "Book Value / Share (Assets-Liab)": base.get("book_value_per_share_bs_equity", np.nan),

        "DCF Price (5yr)": base.get("dcf_price", np.nan),
        "Residual Income Price": base.get("residual_income_price", np.nan),
        "Asset Based Price": base.get("asset_based_price", np.nan),
        "SOTP Price": base.get("sotp_price", np.nan),
        "Dividend Discount Price": base.get("ddm_price", np.nan),
        "Earnings Power Value (EPV) Price": base.get("epv_price", np.nan),
        "Option Pricing Value": base.get("option_value", np.nan),

        "DCF Estimate": base.get("dcf_estimate", ""),

        "DCF Premium/(Discount)": base.get("dcf_prem", np.nan),
        "Residual Income Premium/(Discount)": base.get("ri_prem", np.nan),
        "Asset Based Premium/(Discount)": base.get("asset_prem", np.nan),
        "SOTP Premium/(Discount)": base.get("sotp_prem", np.nan),
        "Dividend Discount Premium/(Discount)": base.get("ddm_prem", np.nan),
        "EPV Premium/(Discount)": base.get("epv_prem", np.nan),
        "Option Pricing Premium/(Discount)": base.get("opt_prem", np.nan),

        "Undervalued Methods Count": base.get("undervalued_methods_count", np.nan),
        "Margin of Safety": base.get("margin_of_safety", np.nan),
        "MOS Buy Price": base.get("mos_buy_price", np.nan),

        "FCF (Yahoo)": base.get("freeCashflow", np.nan),
        "Revenue Growth (Yahoo)": base.get("revenueGrowth", np.nan),
        "Growth Input Used": base.get("growth_input_used", np.nan),
        "Beta": base.get("beta", np.nan),
        "Risk Free Rate": ASSUMPTIONS["risk_free_rate"],
        "Market Risk Premium": ASSUMPTIONS["market_risk_premium"],
        "Cost of Equity": base.get("cost_of_equity", np.nan),
        "WACC": base.get("wacc", np.nan),
        "Terminal Growth Rate": ASSUMPTIONS["terminal_growth_rate"],
        "Valuation Period (yrs)": ASSUMPTIONS["valuation_period"],
        "Tax Rate": ASSUMPTIONS["tax_rate"],

        "FCF Yield": base.get("fcf_yield", np.nan),
        "Trailing PE": base.get("trailing_pe", np.nan),
        "Forward PE": base.get("forward_pe", np.nan),
        "P/B": base.get("price_to_book", np.nan),
        "EV/EBITDA": base.get("ev_to_ebitda", np.nan),
        "Net Debt/EBITDA": base.get("net_debt_to_ebitda", np.nan),

        "ROA": base.get("roa", np.nan),
        "ROE": base.get("roe", np.nan),
        "Gross Margin": base.get("gross_margin", np.nan),
        "Operating Margin": base.get("operating_margin", np.nan),
        "Profit Margin": base.get("profit_margin", np.nan),

        "Debt/Equity": base.get("debt_to_equity", np.nan),
        "Current Ratio": base.get("current_ratio", np.nan),
        "Quick Ratio": base.get("quick_ratio", np.nan),

        "Held % Insiders": base.get("held_pct_insiders", np.nan),
        "Held % Institutions": base.get("held_pct_institutions", np.nan),
        "Short % Float": base.get("short_pct_float", np.nan),
        "Dividend Rate (Yahoo)": base.get("dividend_rate", np.nan),
        "Dividend Yield (Yahoo)": norm_yield(base.get("dividend_yield", np.nan)),

        # Raw Yahoo dividend metadata (as available)
        "Payout Ratio (Yahoo)": base.get("payoutRatio", np.nan),
        "5Y Avg Dividend Yield (Yahoo)": norm_yield(base.get("fiveYearAvgDividendYield", np.nan)),
        "Ex-Dividend Date (Yahoo)": epoch_to_date_str(base.get("exDividendDate", np.nan)),
        "Last Dividend Value (Yahoo)": base.get("lastDividendValue", np.nan),
        "Last Dividend Date (Yahoo)": epoch_to_date_str(base.get("lastDividendDate", np.nan)),

        # Requested calc: trailing annual dividend rate / latest share price
        "Dividend Yield (Latest, Calc)": (
            float(base.get("dividend_rate", np.nan)) / float(base["currentPrice"])
            if (not np.isnan(base.get("dividend_rate", np.nan)) and not np.isnan(base["currentPrice"]) and float(base["currentPrice"]) > 0)
            else np.nan
        ),

        # Delta between Yahoo's dividendYield and the calculated yield above (relative % change)
        "Dividend Yield Δ% (Yahoo→Calc)": (
            ( (float(base.get("dividend_rate", np.nan)) / float(base["currentPrice"])) - float(norm_yield(base.get("dividend_yield", np.nan))) )
            / float(norm_yield(base.get("dividend_yield", np.nan)))
            if (not np.isnan(base.get("dividend_rate", np.nan)) and not np.isnan(base["currentPrice"]) and float(base["currentPrice"]) > 0
                and not np.isnan(base.get("dividend_yield", np.nan)) and float(norm_yield(base.get("dividend_yield", np.nan))) != 0)
            else np.nan
        ),


        "Reverse DCF Implied Growth": base.get("reverse_dcf_implied_growth", np.nan),
        "PEG Ratio": base.get("peg_ratio", np.nan),

        "Enterprise Value (Yahoo/Calc)": base.get("enterprise_value_yahoo_or_calc", np.nan),
    }

    # Add technical columns with the exact names the screener expects
    out_row.update(tech)

    # Add a simple completeness score for quick filtering
    out_row["Data Quality Score"] = data_quality_score(out_row)

    return out_row


def base_from_prev_row(prev: Dict[str, Any], company: UniverseRow) -> Dict[str, Any]:
    """Rebuild the internal base dict (the input schema for calc_valuations) from a previously exported UI row."""
    def g(k: str, default=np.nan):
        v = prev.get(k, default)
        return default if v is None else v

    base: Dict[str, Any] = {
        "ticker": normalize_ticker(str(g("Ticker", ""))),
        "company_name": str(g("Company", "")) or company.name,
        "sector": str(g("Sector", "")) or company.sector,
        "industry": str(g("Industry", "")) or company.industry,
        "currency": str(g("Currency", "AUD")) or "AUD",
        "asof": datetime.now().strftime("%Y-%m-%d"),

        "currentPrice": g("Price", np.nan),
        "marketCap": g("Market Cap", np.nan),
        "sharesOutstanding": g("Shares Out", np.nan),

        "cash": g("Cash", np.nan),
        "totalDebt": g("Total Debt", np.nan),
        "total_assets": g("Total Assets", np.nan),
        "total_liabilities": g("Total Liabilities", np.nan),
        "net_tangible_assets": g("NTA", np.nan),

        # Yahoo-like fundamental inputs
        "freeCashflow": g("FCF (Yahoo)", np.nan),
        "revenueGrowth": g("Revenue Growth (Yahoo)", np.nan),
        "earningsGrowth": g("Earnings Growth (Yahoo)", np.nan),
        "beta": g("Beta", np.nan),

        # Per-share + profitability
        "bookValue": g("Book Value / Share (Yahoo)", np.nan),
        "returnOnEquity": g("ROE", np.nan),
        "trailingPE": g("Trailing PE", np.nan),
        "enterpriseToEbitda": g("EV/EBITDA", np.nan),
        "priceToBook": g("P/B", np.nan),
        "totalRevenue": g("Revenue (Yahoo)", np.nan),
        "ebitda": g("EBITDA (Yahoo)", np.nan),
        "operatingMargins": g("Operating Margin", np.nan),

        # Dividends
        "dividendRate": g("Dividend Rate (Yahoo)", np.nan),
        "dividendYield": g("Dividend Yield (Yahoo)", np.nan),
        "payoutRatio": g("Payout Ratio (Yahoo)", np.nan),
        "fiveYearAvgDividendYield": g("5Y Avg Dividend Yield (Yahoo)", np.nan),
        "exDividendDate": g("Ex-Dividend Date (Yahoo)", np.nan),
        "lastDividendValue": g("Last Dividend Value (Yahoo)", np.nan),
        "lastDividendDate": g("Last Dividend Date (Yahoo)", np.nan),

        # Helpful derived fields (if present)
        "net_debt": g("Net Debt (Debt - Cash)", np.nan),
        "nta_per_share": g("NTA / Share", np.nan),
        "book_value_total_yahoo": g("Book Value (Total, Yahoo)", np.nan),
        "book_value_total_assets_liab": g("Book Value (Total, Assets-Liab)", np.nan),
        "enterprise_value_yahoo_or_calc": g("Enterprise Value (Yahoo/Calc)", np.nan),
        "net_debt_to_ebitda": g("Net Debt/EBITDA", np.nan),

        # Optional: insiders / inst / short interest (already in UI row)
        "held_pct_insiders": g("Held % Insiders", np.nan),
        "held_pct_institutions": g("Held % Institutions", np.nan),
        "short_pct_float": g("Short % Float", np.nan),
        "profit_margin": g("Profit Margin", np.nan),
        "gross_margin": g("Gross Margin", np.nan),
    }
    return base


def fetch_one_technicals(
    prev_row: Dict[str, Any],
    company: UniverseRow,
    benchmark_rets: Optional[pd.Series],
    history_period: str,
    history_interval: str,
    include_long_tf: bool = True,
    hist_override: Optional[pd.DataFrame] = None,
) -> Optional[Dict[str, Any]]:
    """
    Technical-only refresh (fast path).

    IMPORTANT: This function MUST NOT "rebuild" the row schema from scratch, otherwise we risk
    wiping fundamentals/valuation fields during intraday runs (which makes most UI charts blank).

    Strategy:
      - Start from the previously exported UI row (prev_row) and preserve all fundamental fields.
      - Refresh: Price/Market Cap (if possible) + technical indicators.
      - Recompute only *price-sensitive* derived fields (prem/discount, FCF yield, dividend calc yields, EV, etc.)
        using the preserved baseline fields from the weekly fundamentals run.
    """
    if not isinstance(prev_row, dict):
        return None

    ticker_code = normalize_ticker(str(prev_row.get("Ticker", "")))
    if not ticker_code:
        return None

    # Copy UI row (preserve schema/fundamentals)
    row: Dict[str, Any] = dict(prev_row)

    # Ensure identity fields are present / up to date
    row["Ticker"] = ticker_code
    if company is not None:
        row["Company"] = row.get("Company") or company.name
        row["Sector"] = row.get("Sector") or company.sector
        row["Industry"] = row.get("Industry") or company.industry
    row["As Of"] = datetime.now().strftime("%Y-%m-%d")

    # Deterministic GICS rollup (stable across runs; used for RRG → screener filtering)
    try:
        g = normalize_gics_sector(row.get("Sector", ""), row.get("Industry", ""))
        if not row.get("GICS Sector") or str(row.get("GICS Sector")).strip() in ("", "Unknown"):
            row["GICS Sector"] = g
    except Exception:
        if not row.get("GICS Sector"):
            row["GICS Sector"] = "Unknown"

    def _f(x):
        try:
            if x is None:
                return np.nan
            v = float(x)
            return v
        except Exception:
            return np.nan

    def _prem(fair_price, px):
        fair = _f(fair_price)
        px2 = _f(px)
        if not np.isfinite(fair) or not np.isfinite(px2) or px2 <= 0:
            return np.nan
        return float((fair - px2) / px2)

    # Fetch history (prefer bulk override for speed)
    sym = yahoo_symbol(ticker_code)
    hist = hist_override if (isinstance(hist_override, pd.DataFrame) and not hist_override.empty) else fetch_history(sym, history_period, history_interval)

    # Update current price from history (Close)
    px = _f(row.get("Price"))
    if hist is not None and not hist.empty and "Close" in hist.columns:
        try:
            px_new = float(pd.to_numeric(hist["Close"], errors="coerce").dropna().iloc[-1])
            if np.isfinite(px_new) and px_new > 0:
                px = px_new
                row["Price"] = px
        except Exception:
            pass

    # Long timeframes only if requested
    hist_w = None
    hist_m = None
    if include_long_tf:
        try:
            hist_w = fetch_history(sym, "5y", "1wk")
        except Exception:
            hist_w = None
        try:
            hist_m = fetch_history(sym, "10y", "1mo")
        except Exception:
            hist_m = None

    # Compute technicals (this returns columns with the exact UI names)
    tech = compute_technicals(hist, benchmark_rets=benchmark_rets, weekly_hist=hist_w, monthly_hist=hist_m)
    # Only overwrite fields when we actually computed a finite value.
    # This prevents intraday runs with --no_long_tf from wiping weekly/monthly columns (SMA200W, W/M S/R, etc.).
    for k, v in tech.items():
        try:
            if v is None:
                continue
            if isinstance(v, (int, float, np.floating)) and (not np.isfinite(float(v))):
                continue
        except Exception:
            pass
        row[k] = v

    # ---- Price-sensitive derived fields (keep fundamentals, update deltas) ----

    # Market Cap: if shares outstanding exists, recompute from fresh price
    sh = _f(row.get("Shares Out"))
    if np.isfinite(px) and px > 0 and np.isfinite(sh) and sh > 0:
        row["Market Cap"] = float(sh * px)

    # Net debt + EV (pure arithmetic; keeps downstream charts stable)
    cash = _f(row.get("Cash"))
    debt = _f(row.get("Total Debt"))
    if np.isfinite(debt) or np.isfinite(cash):
        if not np.isfinite(debt):
            debt = 0.0
        if not np.isfinite(cash):
            cash = 0.0
        row["Net Debt (Debt - Cash)"] = float(debt - cash)

    mcap = _f(row.get("Market Cap"))
    if np.isfinite(mcap):
        ev = mcap
        if np.isfinite(debt):
            ev += debt
        if np.isfinite(cash):
            ev -= cash
        row["Enterprise Value (Yahoo/Calc)"] = float(ev)

    # FCF Yield = FCF / Market Cap
    fcf = _f(row.get("FCF (Yahoo)"))
    if np.isfinite(fcf) and np.isfinite(mcap) and mcap > 0:
        row["FCF Yield"] = float(fcf / mcap)

    # Recompute valuation premiums using preserved fair-value columns + updated price
    row["DCF Premium/(Discount)"] = _prem(row.get("DCF Price (5yr)"), px)
    row["Residual Income Premium/(Discount)"] = _prem(row.get("Residual Income Price"), px)
    row["Asset Based Premium/(Discount)"] = _prem(row.get("Asset Based Price"), px)
    row["SOTP Premium/(Discount)"] = _prem(row.get("SOTP Price"), px)
    row["Dividend Discount Premium/(Discount)"] = _prem(row.get("Dividend Discount Price"), px)
    row["EPV Premium/(Discount)"] = _prem(row.get("Earnings Power Value (EPV) Price"), px)
    row["Option Pricing Premium/(Discount)"] = _prem(row.get("Option Pricing Value"), px)

    # Undervalued methods count (same logic as calc_valuations)
    prem_keys = [
        "DCF Premium/(Discount)",
        "Residual Income Premium/(Discount)",
        "Asset Based Premium/(Discount)",
        "SOTP Premium/(Discount)",
        "Dividend Discount Premium/(Discount)",
    ]
    try:
        row["Undervalued Methods Count"] = int(sum(1 for k in prem_keys if np.isfinite(_f(row.get(k))) and _f(row.get(k)) > 0))
    except Exception:
        pass

    # Dividend yield derived from dividend rate / current price
    div_rate = _f(row.get("Dividend Rate (Yahoo)"))
    if np.isfinite(div_rate) and np.isfinite(px) and px > 0:
        calc_y = float(div_rate / px)
        row["Dividend Yield (Latest, Calc)"] = calc_y
        # Delta % (Yahoo -> Calc) if Yahoo yield is present
        y_yahoo = _f(norm_yield(row.get("Dividend Yield (Yahoo)")))
        if np.isfinite(y_yahoo) and y_yahoo != 0:
            row["Dividend Yield Δ% (Yahoo→Calc)"] = float((calc_y - y_yahoo) / y_yahoo)

    return row



# --------------------------------------------------------------------------------------
# RRG (Relative Rotation Graph) exports (sector indices via Yahoo)
# - Ports the tilescreener reference behavior:
#   - Algo: JdK + Simple
#   - TF: daily/weekly/monthly (UI-selectable; we export a bundle)
#   - Tail controlled in UI (we export longer trails)
# - Uses yf.download in bulk for robustness/speed (12 tickers total)
# --------------------------------------------------------------------------------------


RRG_SECTORS = [
    {"symbol": "^AXTJ", "name": "Communication Services"},
    {"symbol": "^AXDJ", "name": "Consumer Discretionary"},
    {"symbol": "^AXSJ", "name": "Consumer Staples"},
    {"symbol": "^AXEJ", "name": "Energy"},
    {"symbol": "^AXFJ", "name": "Financials"},
    {"symbol": "^AXHJ", "name": "Health Care"},
    {"symbol": "^AXNJ", "name": "Industrials"},
    {"symbol": "^AXIJ", "name": "Information Technology"},
    {"symbol": "^AXMJ", "name": "Materials"},
    {"symbol": "^AXRE", "name": "Real Estate"},
    {"symbol": "^AXUJ", "name": "Utilities"},
]


def ema_series(values: List[float], period: int) -> List[float]:
    if not values or period <= 1:
        return list(values)
    alpha = 2 / (period + 1)
    out: List[float] = []
    prev = float(values[0])
    out.append(prev)
    for v in values[1:]:
        prev = alpha * float(v) + (1 - alpha) * prev
        out.append(prev)
    return out


def _to_day_key_ms(t_ms: int) -> str:
    return datetime.fromtimestamp(int(t_ms) / 1000, tz=ZoneInfo("UTC")).date().isoformat()


def rrg_series_simple(series: List[Dict[str, Any]], bench: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """A simple RRG approximation (ported from tilescreener behavior).

    RS = sector/bench
    RS-Ratio = 100 * RS / SMA10(RS)
    RS-Mom   = 100 + %chg(RS over 10 bars)
    """
    map_b = {_to_day_key_ms(int(p["t"])): p["v"] for p in bench if p.get("v") is not None}
    rs: List[Dict[str, Any]] = []
    for p in series:
        if p.get("v") is None:
            continue
        key = _to_day_key_ms(int(p["t"]))
        bv = map_b.get(key)
        if bv is None or bv == 0:
            continue
        rs.append({"t": int(p["t"]), "v": float(p["v"]) / float(bv)})

    if len(rs) < 15:
        return []

    trail: List[Dict[str, Any]] = []
    for i in range(10, len(rs)):
        window = [x["v"] for x in rs[max(0, i - 10): i + 1]]
        avg = sum(window) / len(window) if window else rs[i]["v"]
        ratio = (rs[i]["v"] / avg) * 100.0 if avg else 100.0
        prev = rs[i - 10]["v"]
        mom_pct = ((rs[i]["v"] / prev) - 1.0) * 100.0 if prev else 0.0
        momentum = 100.0 + mom_pct
        trail.append({"t": rs[i]["t"], "x": float(ratio), "y": float(momentum)})

    return trail


def rrg_series_jdk(
    series: List[Dict[str, Any]],
    bench: List[Dict[str, Any]],
    short: int = 10,
    long: int = 30,
    momN: int = 10,
) -> List[Dict[str, Any]]:
    """JdK-style RS-Ratio / RS-Momentum.

    RS = sector/bench
    RS-Ratio = 100 * EMA_short(RS) / EMA_long(RS)
    RS-Mom   = 100 + 100 * ((RS-Ratio / RS-Ratio[momN]) - 1)
    """
    short = max(2, min(60, int(short)))
    long = max(short + 1, min(120, int(long)))
    momN = max(2, min(60, int(momN)))

    map_b = {_to_day_key_ms(int(p["t"])): p["v"] for p in bench if p.get("v") is not None}

    ts: List[int] = []
    rs_vals: List[float] = []
    for p in series:
        if p.get("v") is None:
            continue
        key = _to_day_key_ms(int(p["t"]))
        bv = map_b.get(key)
        if bv is None or bv == 0:
            continue
        ts.append(int(p["t"]))
        rs_vals.append(float(p["v"]) / float(bv))

    if len(rs_vals) < long + momN + 2:
        return []

    rs_ema_s = ema_series(rs_vals, short)
    rs_ema_l = ema_series(rs_vals, long)
    rs_ratio = [100.0 * (rs_ema_s[i] / (rs_ema_l[i] if rs_ema_l[i] else rs_ema_s[i])) for i in range(len(rs_vals))]

    trail: List[Dict[str, Any]] = []
    for i in range(momN, len(rs_ratio)):
        x = float(rs_ratio[i])
        prev = float(rs_ratio[i - momN])
        y = 100.0 + 100.0 * ((x / prev) - 1.0) if prev else 100.0
        trail.append({"t": ts[i], "x": x, "y": float(y)})

    return trail


def _bulk_fetch_close_series(symbols: List[str], period: str, interval: str) -> Dict[str, List[Dict[str, Any]]]:
    """Bulk fetch close series via yf.download. Returns {symbol: [{t(ms), v(close)}...]}."""
    out: Dict[str, List[Dict[str, Any]]] = {}
    try:
        df = yf.download(
            tickers=" ".join(symbols),
            period=period,
            interval=interval,
            auto_adjust=True,
            progress=False,
            group_by="ticker",
            threads=True,
        )
    except Exception:
        return out

    frames = _split_download_frame(df, symbols)
    for sym, fr in frames.items():
        if fr is None or fr.empty or "Close" not in fr.columns:
            continue
        close = pd.to_numeric(fr["Close"], errors="coerce").dropna()
        pts: List[Dict[str, Any]] = []
        for ts, v in close.items():
            try:
                t_ms = int(pd.Timestamp(ts).to_pydatetime().replace(tzinfo=ZoneInfo("UTC")).timestamp() * 1000)
                pts.append({"t": t_ms, "v": float(v)})
            except Exception:
                continue
        pts.sort(key=lambda p: p["t"])
        if pts:
            out[sym] = pts
    return out


def write_rrg_sectors(public_dir: Path, tf: str = "all", bench: str = "^AXJO") -> Optional[Path]:
    """Write public_dir/rrg_sectors.json, bundling TFs and algos.

    JSON shape:
      {
        at, bench, sectors: [...],
        data: {
          weekly: {meta, jdk:{short,long,momN,series:[...]}, simple:{series:[...]}},
          ...
        }
      }
    """
    tf = (tf or "all").lower()
    tfs = ["daily", "weekly", "monthly"] if tf == "all" else [tf]
    tfs = [t for t in tfs if t in ("daily", "weekly", "monthly")]
    if not tfs:
        tfs = ["weekly"]

    symbols = [bench] + [s["symbol"] for s in RRG_SECTORS]
    payload: Dict[str, Any] = {
        "at": datetime.now(ZoneInfo("UTC")).isoformat(timespec="seconds"),
        "bench": bench,
        "sectors": RRG_SECTORS,
        "data": {},
    }

    for tf_ in tfs:
        period = "6mo" if tf_ == "daily" else ("2y" if tf_ == "weekly" else "5y")
        interval = "1d" if tf_ == "daily" else ("1wk" if tf_ == "weekly" else "1mo")

        ser_map = _bulk_fetch_close_series(symbols, period=period, interval=interval)
        bench_series = ser_map.get(bench, [])

        payload["data"][tf_] = {"meta": {"period": period, "interval": interval}}

        if not bench_series:
            payload["data"][tf_]["error"] = f"Missing benchmark history for {bench} (period={period}, interval={interval})"
            payload["data"][tf_]["jdk"] = {"series": []}
            payload["data"][tf_]["simple"] = {"series": []}
            continue

        # Defaults aligned to typical JdK usage; tweaked per TF for stability
        if tf_ == "daily":
            short, long, momN = 7, 14, 7
        elif tf_ == "weekly":
            short, long, momN = 10, 30, 10
        else:
            short, long, momN = 5, 10, 5

        out_jdk: List[Dict[str, Any]] = []
        out_simple: List[Dict[str, Any]] = []

        for s in RRG_SECTORS:
            sec_series = ser_map.get(s["symbol"], [])
            if not sec_series:
                out_jdk.append({"symbol": s["symbol"], "name": s["name"], "latest": None, "trail": []})
                out_simple.append({"symbol": s["symbol"], "name": s["name"], "latest": None, "trail": []})
                continue

            trail_jdk = rrg_series_jdk(sec_series, bench_series, short=short, long=long, momN=momN)
            trail_simple = rrg_series_simple(sec_series, bench_series)

            # Keep file size reasonable; UI controls tail anyway
            trail_jdk = trail_jdk[-80:]
            trail_simple = trail_simple[-80:]

            last_jdk = trail_jdk[-1] if trail_jdk else None
            last_simple = trail_simple[-1] if trail_simple else None

            out_jdk.append({"symbol": s["symbol"], "name": s["name"], "latest": last_jdk, "trail": trail_jdk})
            out_simple.append({"symbol": s["symbol"], "name": s["name"], "latest": last_simple, "trail": trail_simple})

        payload["data"][tf_]["jdk"] = {"short": short, "long": long, "momN": momN, "series": out_jdk}
        payload["data"][tf_]["simple"] = {"series": out_simple}

    out_path = public_dir / "rrg_sectors.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return out_path

def main() -> None:
    ap = argparse.ArgumentParser(description="Compute intrinsic values + technicals for all ASX tickers and export to Excel/CSV/JSON.")
    ap.add_argument("--scraper", default="scrape_asx_universe.py", help="Path to scrape_asx_universe.py")
    ap.add_argument("--universe_csv", default="data/universe/asx_universe.csv", help="Universe CSV output (cached)")
    ap.add_argument("--tickers_txt", default="", help="Optional tickers TXT output (cached)")
    ap.add_argument("--universe_max_age_days", type=int, default=30, help="Refresh universe only if older than this many days")
    ap.add_argument("--universe_source", default="auto", choices=["auto", "asx_csv", "marketindex"], help="Universe source")
    ap.add_argument("--universe_sleep", type=float, default=0.5, help="Sleep for fallback crawling (universe script)")
    ap.add_argument("--materials_energy_csv", default="data/roster/asx_materials_energy_companies.csv", help="Optional join (products/stage) if exists")
    ap.add_argument("--output_xlsx", default="", help="Output Excel path (default: data/valuations/...)")
    ap.add_argument("--public_dir", default="public/data", help="Write latest.* outputs into this directory for Cloudflare Pages")
    ap.add_argument("--sleep", type=float, default=0.25, help="Sleep seconds between Yahoo calls")
    ap.add_argument("--limit", type=int, default=0, help="Limit tickers for testing (0 = no limit)")
    ap.add_argument("--resume", action="store_true", help="Resume from a progress CSV if present")
    ap.add_argument("--progress_csv", default="data/valuations/_progress_intrinsic.csv", help="Progress CSV for --resume")
    ap.add_argument("--benchmark", default="^AXJO", help="Benchmark ticker for beta calc (default ^AXJO)")
    ap.add_argument("--history_period", default="1y", help="yfinance history period (default 1y)")
    ap.add_argument("--history_interval", default="1d", help="yfinance history interval (default 1d)")
    ap.add_argument(
        "--bulk_prices",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Bulk-download daily OHLCV in chunks via yfinance.download (much faster than per-ticker history)",
    )
    ap.add_argument("--bulk_chunk_size", type=int, default=250, help="Tickers per yfinance.download request when --bulk_prices is enabled")
    ap.add_argument("--disable_technicals", action="store_true", help="Disable technical calculations (faster)")
    ap.add_argument("--update_mode", default="full", choices=["full", "technicals"], help="full = refresh fundamentals+technicals; technicals = refresh only technicals/price using prior output")
    ap.add_argument("--no_long_tf", action="store_true", help="Skip weekly/monthly technical overlays (faster)")
    ap.add_argument("--generate_rrg", action="store_true", help="Write RRG sector dataset to public_dir (rrg_sectors.json)")
    ap.add_argument("--rrg_tf", default="all", choices=["daily","weekly","monthly","all"], help="RRG timeframe bundle to export (default all)")
    args = ap.parse_args()

    scraper_path = Path(args.scraper).resolve()
    universe_csv = Path(args.universe_csv).resolve()
    tickers_txt = Path(args.tickers_txt).resolve() if args.tickers_txt else None

    maybe_refresh_universe(
        scraper_path=scraper_path,
        universe_csv=universe_csv,
        tickers_txt=tickers_txt,
        max_age_days=args.universe_max_age_days,
        source=args.universe_source,
        sleep_s=args.universe_sleep,
    )

    universe = load_universe(universe_csv)
    if args.limit and args.limit > 0:
        universe = universe[: args.limit]
    print(f"[info] tickers to process: {len(universe)}")

    me_df = maybe_load_materials_energy(Path(args.materials_energy_csv).resolve())

    # ----------------------------------------------------------------------------------
    # Bulk price history (fast path)
    # ----------------------------------------------------------------------------------
    daily_hist_map: Dict[str, pd.DataFrame] = {}
    if not args.disable_technicals and bool(args.bulk_prices):
        syms = [yahoo_symbol(u.ticker) for u in universe]
        if args.benchmark:
            syms = syms + [str(args.benchmark)]
        print(f"[info] bulk downloading price history in chunks (n={len(syms)}; chunk={int(args.bulk_chunk_size)})")
        daily_hist_map = bulk_fetch_history(syms, args.history_period, args.history_interval, chunk_size=int(args.bulk_chunk_size))

    # Benchmark returns (daily)
    benchmark_rets = None
    if not args.disable_technicals:
        try:
            b_hist = daily_hist_map.get(str(args.benchmark), pd.DataFrame())
            if b_hist.empty:
                b_hist = fetch_history(args.benchmark, args.history_period, args.history_interval)
            if not b_hist.empty and "Close" in b_hist.columns:
                benchmark_rets = b_hist["Close"].astype(float).pct_change().dropna()
        except Exception:
            benchmark_rets = None


    progress_csv = Path(args.progress_csv).resolve()
    processed = set()
    rows_out: List[Dict[str, Any]] = []

    # Load previous rows for resume / technical-only update mode
    prev_map: Dict[str, Dict[str, Any]] = {}
    prev_df: Optional[pd.DataFrame] = None

    def _load_prev() -> None:
        nonlocal prev_map, prev_df
        if progress_csv.exists():
            try:
                prev_df = pd.read_csv(progress_csv)
            except Exception:
                prev_df = None
        if prev_df is None:
            # fallback to last published dataset (best effort)
            fallback_web = Path(args.public_dir).resolve() / "latest_web.json"
            if fallback_web.exists():
                try:
                    prev_df = pd.read_json(fallback_web)
                except Exception:
                    prev_df = None
            fallback_json = Path(args.public_dir).resolve() / "latest.json"
            if prev_df is None and fallback_json.exists():
                try:
                    prev_df = pd.read_json(fallback_json)
                except Exception:
                    prev_df = None
            fallback_csv = Path(args.public_dir).resolve() / "latest.csv"

            if prev_df is None and fallback_csv.exists():
                try:
                    prev_df = pd.read_csv(fallback_csv)
                except Exception:
                    prev_df = None
        if prev_df is not None and "Ticker" in prev_df.columns:
            prev_map = {normalize_ticker(str(r["Ticker"])): r for r in prev_df.to_dict(orient="records")}

    if args.resume or args.update_mode == "technicals":
        _load_prev()

    if args.resume and args.update_mode != "technicals" and progress_csv.exists() and prev_df is not None and "Ticker" in prev_df.columns:
        processed = set(prev_df["Ticker"].astype(str).map(normalize_ticker).tolist())
        rows_out = prev_df.to_dict(orient="records")
        print(f"[resume] loaded {len(processed)} processed tickers from {progress_csv}")

    include_long_tf = not bool(args.no_long_tf)

    t0 = time.time()

    if args.update_mode == "technicals":
        # Technical-only refresh: rebuild rows from previous output, updating price/technicals and re-running scoring.
        if not prev_map:
            print("[warn] update_mode=technicals but no previous dataset found; falling back to full refresh.")
        else:
            print(f"[mode] technicals-only refresh using prior rows: {len(prev_map)}")

    for i, u in enumerate(universe, start=1):
        # In technical-only mode we still iterate the universe so new tickers get pulled in.
        prev_row = prev_map.get(u.ticker)

        sym_u = yahoo_symbol(u.ticker)
        hist_override_u = daily_hist_map.get(sym_u) if daily_hist_map else None

        if args.update_mode == "technicals" and prev_row is not None:
            print(f"[{i}/{len(universe)}] {u.ticker} (technicals)")
            try:
                r = fetch_one_technicals(prev_row, u, benchmark_rets=benchmark_rets, history_period=args.history_period, history_interval=args.history_interval, include_long_tf=include_long_tf, hist_override=hist_override_u)
                if r is not None:
                    rows_out.append(r)
                    processed.add(u.ticker)
            except Exception as e:
                print(f"[warn] {u.ticker} technicals update failed; keeping prior row: {e}")
                rows_out.append(prev_row)
                processed.add(u.ticker)
        else:
            if u.ticker in processed and args.update_mode != "technicals":
                continue
            print(f"[{i}/{len(universe)}] {u.ticker}")
            try:
                r = fetch_one(u.ticker, u, benchmark_rets=benchmark_rets, history_period=args.history_period, history_interval=args.history_interval, disable_technicals=args.disable_technicals, include_long_tf=include_long_tf, hist_override=hist_override_u)
                if r is not None:
                    rows_out.append(r)
                    processed.add(u.ticker)
            except Exception as e:
                print(f"[warn] {u.ticker} failed: {e}")

        if args.resume and args.update_mode != "technicals" and (len(processed) % 50 == 0):
            ensure_parent_dir(progress_csv)
            pd.DataFrame(rows_out).to_csv(progress_csv, index=False)

        sleep_s = max(0.0, float(args.sleep))
        if sleep_s > 0:
            # In technical-only mode, with bulk daily prices and no weekly/monthly overlays,
            # we can skip per-ticker sleeping (no per-ticker HTTP calls are expected).
            fast_technicals = (
                args.update_mode == "technicals"
                and prev_row is not None
                and bool(args.bulk_prices)
                and bool(args.no_long_tf)
                and isinstance(hist_override_u, pd.DataFrame)
                and not hist_override_u.empty
            )
            if not fast_technicals:
                time.sleep(sleep_s)
    dt = time.time() - t0
    print(f"[done] processed={len(processed)} rows in {dt/60.0:.1f} min")

    if not rows_out:
        print("[fatal] no data collected; aborting.")
        sys.exit(2)

    df = pd.DataFrame(rows_out)

    # Optional join: materials/energy products/stage
    if me_df is not None:
        df["_ticker_norm"] = df["Ticker"].astype(str).map(normalize_ticker)
        me_df2 = me_df.copy()
        me_df2["_ticker_norm"] = me_df2["ticker"].astype(str).map(normalize_ticker)
        df = df.merge(me_df2[["_ticker_norm"] + [c for c in ["products", "stage"] if c in me_df2.columns]],
                      on="_ticker_norm", how="left")
        if "products" in df.columns:
            df.rename(columns={"products": "Products (Materials/Energy)", "stage": "Stage (Materials/Energy)"}, inplace=True)
        df.drop(columns=["_ticker_norm"], inplace=True)

    
    # ----------------------------------------------------------------------------------
    # Screener scoring (0-100): Value + Quality + Risk (+ liquidity)
    # ----------------------------------------------------------------------------------
    def _pct_rank(s: pd.Series, ascending: bool = True) -> pd.Series:
        s2 = pd.to_numeric(s, errors="coerce")
        return s2.rank(pct=True, ascending=ascending)

    # Value: DCF discount, FCF yield, MOS upside, low P/B
    df["_v_dcf"] = _pct_rank(df.get("DCF Premium/(Discount)", pd.Series(dtype=float)), ascending=True)
    df["_v_fcf"] = _pct_rank(df.get("FCF Yield", pd.Series(dtype=float)), ascending=True)
    mos_up = np.where(
        (pd.to_numeric(df.get("Price", np.nan), errors="coerce") > 0) & pd.notnull(df.get("MOS Buy Price")),
        (pd.to_numeric(df.get("MOS Buy Price"), errors="coerce") - pd.to_numeric(df.get("Price"), errors="coerce")) / pd.to_numeric(df.get("Price"), errors="coerce"),
        np.nan
    )
    df["_v_mos"] = pd.Series(mos_up).rank(pct=True, ascending=True)
    df["_v_pb"]  = 1.0 - _pct_rank(df.get("P/B", pd.Series(dtype=float)), ascending=True)  # lower is better

    value_score = (0.40*df["_v_dcf"] + 0.30*df["_v_fcf"] + 0.20*df["_v_mos"] + 0.10*df["_v_pb"]) * 100.0

    # Quality: ROE, profit margin, low leverage
    df["_q_roe"] = _pct_rank(df.get("ROE", pd.Series(dtype=float)), ascending=True)
    df["_q_pm"]  = _pct_rank(df.get("Profit Margin", pd.Series(dtype=float)), ascending=True)
    df["_q_nd"]  = 1.0 - _pct_rank(df.get("Net Debt/EBITDA", pd.Series(dtype=float)), ascending=True)  # lower is better
    quality_score = (0.55*df["_q_roe"] + 0.30*df["_q_pm"] + 0.15*df["_q_nd"]) * 100.0

    # Risk: lower vol, lower ATR%, smaller drawdowns
    df["_r_vol"] = 1.0 - _pct_rank(df.get("Vol (20d, ann)", pd.Series(dtype=float)), ascending=True)
    df["_r_atr"] = 1.0 - _pct_rank(df.get("ATR% (14)", pd.Series(dtype=float)), ascending=True)
    # Max Drawdown is negative; "less negative" is better.
    df["_r_mdd"] = _pct_rank(df.get("Max Drawdown (1y)", pd.Series(dtype=float)), ascending=True)
    risk_score = (0.45*df["_r_vol"] + 0.25*df["_r_atr"] + 0.30*df["_r_mdd"]) * 100.0

    # Liquidity bonus (avg $ volume)
    df["_liq"] = _pct_rank(df.get("Avg $Vol 20d", pd.Series(dtype=float)), ascending=True)
    liquidity_bonus = df["_liq"] * 10.0  # +0 to +10

    df["Liquidity Bonus"] = liquidity_bonus.round(2)

    df["Value Score"] = value_score.round(2)
    df["Quality Score"] = quality_score.round(2)
    df["Risk Score"] = risk_score.round(2)
    df["Screener Score"] = (0.45*value_score + 0.30*quality_score + 0.25*risk_score + liquidity_bonus).clip(0,100).round(2)


    # Output path (full workbook)
    if args.output_xlsx:
        out_xlsx = Path(args.output_xlsx).resolve()
    else:
        out_xlsx = Path("data/valuations").resolve() / f"ASX_Intrinsic_Valuations_{now_stamp()}.xlsx"
    ensure_parent_dir(out_xlsx)

    # Write full workbook (Valuations only for speed; you can add extra sheets later)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Valuations", index=False)
        ws = writer.book["Valuations"]
        format_sheet(ws)

    if args.resume:
        ensure_parent_dir(progress_csv)
        df.to_csv(progress_csv, index=False)

    print(f"[ok] wrote: {out_xlsx}")

    # Export web artifacts
    public_dir = Path(args.public_dir).resolve()
    public_dir.mkdir(parents=True, exist_ok=True)

    # Publish a "full" Excel download for humans (kept separate from the screener JSON)
    # Cloudflare Pages has per-file size limits; if the workbook is large we publish a zipped copy instead.
    FULL_XLSX_MAX_BYTES = 24 * 1024 * 1024  # keep a little buffer under 25 MiB
    full_xlsx_name = None
    full_xlsx_zip_name = None
    try:
        size_bytes = out_xlsx.stat().st_size
        if size_bytes <= FULL_XLSX_MAX_BYTES:
            full_target = public_dir / "latest_full.xlsx"
            shutil.copy2(out_xlsx, full_target)
            full_xlsx_name = "latest_full.xlsx"
        else:
            zip_target = public_dir / "latest_full.xlsx.zip"
            with zipfile.ZipFile(zip_target, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                zf.write(out_xlsx, arcname="latest_full.xlsx")
            full_xlsx_zip_name = "latest_full.xlsx.zip"
    except Exception:
        pass

    # Web-friendly subset (keeps latest.xlsx smaller for Cloudflare Pages)
    WEB_COLS = [
        "Ticker","Company","GICS Sector","Sector","Industry",
        "Price","Market Cap",
        "DCF Price (5yr)","DCF Premium/(Discount)","FCF Yield","Undervalued Methods Count",
        "RSI14","ATR% (14)","Vol (20d, ann)","Max Drawdown (1y)",
        "% Dist SMA200D",
        "SMA200W",
        "% Dist SMA200W",
        "MACD (12,26)",
        "MACD Signal (9)",
        "MACD Hist (12,26,9)",
        "ADX14",
        "Stoch %K (14)",
        "Stoch %D (3)",
        "BB %B (20,2)",
        "BB Width (20,2)",
        "Support D1",
        "Support D2",
        "Resistance D1",
        "Resistance D2",
        "Support D1 %",
        "Support D2 %",
        "Resistance D1 %",
        "Resistance D2 %",
        "R:R (D)",
        "Support W1",
        "Support W2",
        "Resistance W1",
        "Resistance W2",
        "Support W1 %",
        "Support W2 %",
        "Resistance W1 %",
        "Resistance W2 %",
        "R:R (W)",
        "Support M1",
        "Support M2",
        "Resistance M1",
        "Resistance M2",
        "Support M1 %",
        "Support M2 %",
        "Resistance M1 %",
        "Resistance M2 %",
        "R:R (M)",
        "Return 1m","Return 3m","Return 12m",
        "% From 52W High","% From 52W Low",
        "Avg $Vol 20d","Avg Vol 20d",
        "Beta vs Benchmark (1y)",
        "ROE","P/B","Net Debt/EBITDA",
        "MOS Buy Price","Margin of Safety",
        "Data Quality Score",
        "As Of",
        "Screener Score",
        "Liquidity Bonus",
        "Value Score",
        "Quality Score",
        "Risk Score",
        "Book Value (Total, Assets-Liab)",
        "Book Value / Share (Assets-Liab)",
        "Book Value / Share (Yahoo)",
        "Profit Margin",
        "Dividend Rate (Yahoo)",
        "Dividend Yield (Yahoo)",
        "Payout Ratio (Yahoo)",
        "5Y Avg Dividend Yield (Yahoo)",
        "Ex-Dividend Date (Yahoo)",
        "Last Dividend Value (Yahoo)",
        "Last Dividend Date (Yahoo)",
        "Dividend Yield (Latest, Calc)",
        "Dividend Yield Δ% (Yahoo→Calc)",
        "Held % Insiders",
        "Held % Institutions"
    ]
    web_df = df[[c for c in WEB_COLS if c in df.columns]].copy()

    # Web JSON (smaller, used by screener UI)
    safe_write_json_records(web_df, public_dir / "latest_web.json")

    latest_xlsx = public_dir / "latest.xlsx"
    with pd.ExcelWriter(latest_xlsx, engine="openpyxl") as writer2:
        web_df.to_excel(writer2, sheet_name="Valuations", index=False)
        ws2 = writer2.book["Valuations"]
        format_sheet(ws2)

    # CSV + JSON (full dataset; UI uses JSON)
    df.to_csv(public_dir / "latest.csv", index=False)
    safe_write_json_records(df, public_dir / "latest.json")

    # Parquet (optional)
    parquet_path = public_dir / "latest.parquet"
    parquet_ok = False
    try:
        df.to_parquet(parquet_path, index=False)
        parquet_ok = True
    except Exception:
        if parquet_path.exists():
            try:
                parquet_path.unlink()
            except Exception:
                pass


    # RRG sector dataset (used by the web UI)
    try:
        if args.generate_rrg:
            write_rrg_sectors(public_dir, tf=args.rrg_tf, bench=args.benchmark)
    except Exception as e:
        print(f"[warn] RRG generation failed: {e}")

    perth = ZoneInfo("Australia/Perth")
    generated_utc = datetime.now(ZoneInfo("UTC")).isoformat(timespec="seconds")
    generated_perth = datetime.now(perth).isoformat(timespec="seconds")

    manifest = {
        "generated_at_utc": generated_utc,
        "generated_at_perth": generated_perth,
        "generated_at_local": datetime.now().isoformat(timespec="seconds"),
        "rows": int(df.shape[0]),
        "columns": int(df.shape[1]),
        "benchmark": args.benchmark,
        "history_period": args.history_period,
        "history_interval": args.history_interval,
        "files": {
            "xlsx": "latest.xlsx",
            "csv": "latest.csv",
            "json": "latest.json",
            "parquet": ("latest.parquet" if parquet_ok else None),
            "rrg_sectors": ("rrg_sectors.json" if (public_dir / "rrg_sectors.json").exists() else None),
        },
    }
    safe_atomic_write(public_dir / "manifest.json", json.dumps(manifest, indent=2))
    print(f"[ok] wrote web exports to: {public_dir}")


if __name__ == "__main__":
    main()
